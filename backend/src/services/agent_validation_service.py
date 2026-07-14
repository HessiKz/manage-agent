"""Post-create AI validation for newly created agents."""

from __future__ import annotations

import asyncio
import io
import json
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from uuid import uuid4

from fastapi import HTTPException
from langchain_core.messages import HumanMessage, SystemMessage
from openpyxl import Workbook, load_workbook
from sqlalchemy import desc, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm.attributes import flag_modified

from src.agents_lib.agent_factory import build_llm
from src.config import settings
from src.core import llm_runtime
from src.core.errors import AppError, ErrorCode
from src.core.file_policy import validate_upload
from src.core.agent_file_roles import is_instruction_file
from src.core.file_text_extract import extract_text
from src.database.session import async_session_maker
from src.models.agent import Agent, AgentStatus
from src.models.agent_action import AgentAction
from src.models.agent_file import AgentFile
from src.models.notification import NotificationSeverity
from src.models.user import User
from src.schemas.agent import AgentInvokeRequest
from src.schemas.agent_action import AgentActionRunRequest
from src.services.agent_action_service import (
    AgentActionService,
    _has_tool_activity,
    _is_advisory_output,
)
from src.services.notification_service import NotificationService
from src.services.agent_execution_service import AgentExecutionService
from src.services.orchestrator_service import OrchestratorService
from src.services.agent_script_service import AgentScriptService

# Ensure demo/custom tools are registered even in non-API entrypoints (scripts/tests).
import src.agents_lib.custom_tools  # noqa: F401


@dataclass
class ValidationFailure:
    phase: str
    message: str
    fixable_in_admin: bool


class AgentValidationService:
    """Runs smoke tests so new agents are runnable before activation."""

    PLANNING_LOCALE = "fa-IR"

    _PLANNING_SYSTEM = (
        "You analyze AI agent configurations before a smoke test for a Persian enterprise platform.\n"
        "Read the config and attached file previews carefully. Identify ambiguity, "
        "missing data, conflicting instructions, non-standard formats, or anything "
        "that could make the agent produce the wrong output.\n"
        "Return ONLY valid JSON with this exact shape:\n"
        '{"analysis": "...", "questions": [{"id": "q1", "text": "...", "context": "...", "options": ["گزینه ۱", "گزینه ۲", "گزینه ۳"]}]}'
        "\nAll user-facing fields MUST be in clear, simple Persian (fa-IR):\n"
        "- analysis: short initial analysis paragraph\n"
        "- questions[].text: the clarifying question\n"
        "- questions[].context: optional short note explaining why this question matters\n"
        "- questions[].options: 3-5 SHORT Persian answer options. The user will pick ONE.\n"
        "  Always include options for EVERY question. Make options concrete and mutually\n"
        "  exclusive single-choice values. The frontend adds a free-text 'سایر' fallback\n"
        "  yourself, so do not include that as an option. Think Cursor-style quick-replies.\n"
        "Generate at most 5 targeted questions. If everything is clear, return questions: []."
    )

    def __init__(self, db: AsyncSession):
        self.db = db
        self.orchestrator = OrchestratorService(db)
        self.actions = AgentActionService(db)
        self.notifications = NotificationService(db)
        self.step_timeout_s = max(30, int(settings.agent_validation_timeout_seconds))
        if llm_runtime.active_provider() == "cursor":
            self.step_timeout_s = max(self.step_timeout_s, 600)
        self.step_retries = 1

    async def validate_after_create(self, agent: Agent, owner: User) -> None:
        failures: list[ValidationFailure] = []
        attached_file: AgentFile | None = None
        agent_id = agent.id

        try:
            fresh = await self.db.get(Agent, agent_id)
            if not fresh or fresh.status == AgentStatus.PAUSED:
                return
            agent = fresh
            validation = dict((agent.config_json or {}).get("validation") or {})
            if validation.get("state") == "done":
                return
            planning = dict(validation.get("planning") or {})
            if planning.get("awaiting_answers") and planning.get("locale") == self.PLANNING_LOCALE:
                return
            if planning and planning.get("locale") != self.PLANNING_LOCALE:
                await self._clear_planning(agent_id)
                planning = {}
                fresh = await self.db.get(Agent, agent_id)
                if not fresh or fresh.status == AgentStatus.PAUSED:
                    return
                agent = fresh
                validation = dict((agent.config_json or {}).get("validation") or {})

            await self._publish_phase(agent_id, "starting")

            await self._publish_phase(agent_id, "instruction_compile")
            try:
                await self._validate_instruction_compile(agent)
            except Exception as exc:  # noqa: BLE001
                failures.append(
                    ValidationFailure(
                        phase="instruction_compile",
                        message=f"{type(exc).__name__}: {exc}",
                        fixable_in_admin=True,
                    )
                )

            await self._publish_phase(agent_id, "tool_resolution")
            try:
                await self._validate_tool_resolution(agent, owner)
            except Exception as exc:  # noqa: BLE001
                failures.append(
                    ValidationFailure(
                        phase="tool_resolution",
                        message=f"{type(exc).__name__}: {exc}",
                        fixable_in_admin=True,
                    )
                )

            # File setup first so planning can inspect a real sample path.
            if (agent.capabilities or {}).get("file_upload_enabled"):
                await self._publish_phase(
                    agent_id,
                    "file_setup",
                    "در حال آماده‌سازی فایل نمونه برای تست…",
                )
                try:
                    attached_file = await self._get_or_create_sample_file(agent)
                except Exception as exc:  # noqa: BLE001
                    failures.append(
                        ValidationFailure(
                            phase="file_setup",
                            message=f"{type(exc).__name__}: {exc}",
                            fixable_in_admin=self._is_fixable(exc),
                        )
                    )

            # Planning may already be done in preflight (before interactive training).
            # Only re-run if we have no analysis yet.
            if not planning.get("analysis"):
                await self._publish_phase(
                    agent_id,
                    "planning",
                    "در حال تحلیل عمیق پیکربندی و فایل‌ها با مدل…",
                )
                try:
                    paused = await self._run_planning_phase(agent, attached_file)
                    if paused:
                        return
                    fresh = await self.db.get(Agent, agent_id)
                    if fresh:
                        agent = fresh
                        planning = dict(
                            ((fresh.config_json or {}).get("validation") or {}).get("planning")
                            or {}
                        )
                except Exception as exc:  # noqa: BLE001
                    failures.append(
                        ValidationFailure(
                            phase="planning",
                            message=f"{type(exc).__name__}: {exc}",
                            fixable_in_admin=self._is_fixable(exc),
                        )
                    )

            if (agent.capabilities or {}).get("file_upload_enabled"):
                script_service = AgentScriptService(self.db)
                if planning.get("answers"):
                    self._stamp_planning_answers(agent, planning)
                    await self.db.flush()
                # Refresh io_schema from sample pair for synth prompt.
                try:
                    pairs = await script_service._sample_pairs(agent)
                    if pairs:
                        from src.services.io_schema_service import build_io_schema_pair, persist_io_schema

                        schema = build_io_schema_pair(pairs[0][0], pairs[0][1])
                        persist_io_schema(agent, schema)
                        await self.db.flush()
                except Exception:  # noqa: BLE001
                    pass

                await self._publish_phase(
                    agent_id,
                    "script_evaluate",
                    "در حال ارزیابی نیاز به اسکریپت و تولید کد پردازش فایل…",
                )
                try:
                    meta = await script_service.generate_if_needed(agent, use_llm=True)
                except Exception as exc:  # noqa: BLE001
                    failures.append(
                        ValidationFailure(
                            phase="script_generate",
                            message=f"{type(exc).__name__}: {exc}",
                            fixable_in_admin=True,
                        )
                    )
                    meta = {}

                if meta.get("needed"):
                    await self._publish_phase(
                        agent_id,
                        "script_verify",
                        "در حال اجرای اسکریپت روی نمونه و مقایسه با خروجی طلایی…",
                    )
                    script_meta = await script_service.verify(agent, use_llm=True)
                    if script_meta.get("last_verify_error"):
                        failures.append(
                            ValidationFailure(
                                phase="script_verify",
                                message=script_meta["last_verify_error"],
                                fixable_in_admin=True,
                            )
                        )

                # Bake time.ir national holidays when relevant — best-effort.
                try:
                    from src.services.holiday_service import build_holiday_calendar, stamp_holiday_calendar
                    table = build_holiday_calendar(agent)
                    if stamp_holiday_calendar(agent, table):
                        flag_modified(agent, "config_json")
                        await self.db.flush()
                except Exception:  # noqa: BLE001
                    pass

            if (agent.capabilities or {}).get("chat_enabled", True):
                await self._publish_phase(
                    agent_id,
                    "invoke",
                    "در حال اجرای تست گفت‌وگوی خودکار با ایجنت…",
                )
                try:
                    await self._run_with_timeout_retry(
                        lambda: self.orchestrator.invoke(
                            agent.id,
                            AgentInvokeRequest(
                                input=self._smoke_prompt(agent, attached_file),
                                thread_id=f"validate-create-{uuid4().hex[:8]}",
                                stream=False,
                            ),
                            owner,
                        )
                    )
                except Exception as exc:  # noqa: BLE001
                    failures.append(
                        ValidationFailure(
                            phase="invoke",
                            message=f"{type(exc).__name__}: {exc}",
                            fixable_in_admin=self._is_fixable(exc),
                        )
                    )

            if (agent.capabilities or {}).get("actions_enabled"):
                action_rows = await self._list_actions(agent.id)
                for action in action_rows:
                    await self._publish_phase(agent_id, f"action:{action.slug}")
                    try:
                        vars_map = self._action_variables(action)
                        if attached_file:
                            vars_map.setdefault("storage_path", attached_file.storage_path)
                        async def _run_action(
                            _slug: str = action.slug,
                            _vars: dict = vars_map,
                            _tool_chain: list = list(action.tool_chain or []),
                        ) -> None:
                            response = await self.actions.run(
                                agent.id,
                                _slug,
                                AgentActionRunRequest(
                                    variables=_vars,
                                    thread_id=f"validate-action-{_slug}-{uuid4().hex[:6]}",
                                ),
                                owner,
                            )
                            if _is_advisory_output(response.output) or (
                                _tool_chain and not _has_tool_activity(response)
                            ):
                                raise ValueError(
                                    (response.output or "Action did not execute tools")[:300]
                                )

                        await self._run_with_timeout_retry(_run_action)
                    except Exception as exc:  # noqa: BLE001
                        failures.append(
                            ValidationFailure(
                                phase=f"action:{action.slug}",
                                message=f"{type(exc).__name__}: {exc}",
                                fixable_in_admin=self._is_fixable(exc),
                            )
                        )

            await self._publish_phase(agent_id, "finishing")
            try:
                fresh_for_guide = await self.db.get(Agent, agent.id)
                if fresh_for_guide:
                    await AgentExecutionService(self.db).build(
                        fresh_for_guide, force_refresh=True
                    )
            except Exception:  # noqa: BLE001
                pass

            report = {
                "validated": True,
                "ok": not failures,
                "state": "done",
                "current_phase": "done",
                "failures": [f.__dict__ for f in failures],
            }
            fresh = await self.db.get(Agent, agent.id)
            if not fresh:
                return
            agent = fresh
            cfg = dict(agent.config_json or {})
            cfg["validation"] = report
            agent.config_json = cfg
            if agent.status == AgentStatus.PAUSED:
                await self.db.commit()
                return
            if not failures:
                agent.status = AgentStatus.ACTIVE
            elif any(f.fixable_in_admin for f in failures):
                agent.status = AgentStatus.ERROR
            else:
                agent.status = AgentStatus.DRAFT

            await self._notify_owner_result(agent, owner, failures)
            if failures:
                await self._notify_admins_if_fixable(agent, failures)

            await self.db.commit()
            await self.db.refresh(agent)
        except Exception:  # noqa: BLE001
            await self.db.rollback()
            raise

    async def _publish_phase(self, agent_id, phase: str, detail: str | None = None) -> None:
        """Expose live validation progress (short-lived session — don't block auth during LLM)."""
        # ponytail: commit main session first — generate_if_needed / file_setup flush() hold
        # a row lock on agent; a second session UPDATE on the same row deadlocks here.
        try:
            await self.db.commit()
        except Exception:  # noqa: BLE001
            await self.db.rollback()
        async with async_session_maker() as db:
            agent = await db.get(Agent, agent_id)
            if not agent or agent.status == AgentStatus.PAUSED:
                return
            cfg = dict(agent.config_json or {})
            validation = dict(cfg.get("validation") or {})
            validation["state"] = "running"
            validation["current_phase"] = phase
            label = detail or phase
            validation["current_detail"] = label
            validation["script_thinking"] = label
            log = list(validation.get("thinking_log") or [])
            log.append(
                {
                    "t": datetime.now(timezone.utc).isoformat(),
                    "phase": phase,
                    "text": label,
                }
            )
            validation["thinking_log"] = log[-50:]
            cfg["validation"] = validation
            agent.config_json = cfg
            try:
                flag_modified(agent, "config_json")
            except Exception:  # noqa: BLE001
                pass
            await db.commit()

    async def _list_actions(self, agent_id) -> list[AgentAction]:
        result = await self.db.execute(
            select(AgentAction).where(AgentAction.agent_id == agent_id).order_by(AgentAction.order_index)
        )
        return list(result.scalars().all())

    async def _get_or_create_sample_file(self, agent: Agent) -> AgentFile:
        result = await self.db.execute(
            select(AgentFile)
            .where(AgentFile.agent_id == agent.id)
            .order_by(desc(AgentFile.created_at))
            .limit(1)
        )
        existing = result.scalar_one_or_none()
        if existing:
            return existing
        return await self._ensure_sample_file(agent)


    async def run_planning_preflight(self, agent: Agent) -> Agent:
        """Run planning Q&A only (before interactive training).

        Order: prepare samples → planning. Does not synthesize scripts or invoke.
        If questions are needed, pauses with awaiting_answers. If not, leaves
        planning.analysis set and does not enter training (caller starts training).
        """
        agent_id = agent.id
        fresh = await self.db.get(Agent, agent_id)
        if not fresh:
            return agent
        agent = fresh
        validation = dict((agent.config_json or {}).get("validation") or {})
        planning = dict(validation.get("planning") or {})
        if planning.get("awaiting_answers") and planning.get("locale") == self.PLANNING_LOCALE:
            return agent
        # Already planned with answers or analysis and no pending questions → done
        if planning.get("analysis") and not planning.get("awaiting_answers"):
            return agent

        await self._publish_phase(
            agent_id,
            "file_setup",
            "در حال آماده‌سازی فایل نمونه برای تحلیل…",
        )
        attached_file = None
        if (agent.capabilities or {}).get("file_upload_enabled"):
            try:
                attached_file = await self._get_or_create_sample_file(agent)
            except Exception:  # noqa: BLE001
                attached_file = None

        await self._publish_phase(
            agent_id,
            "planning",
            "در حال تحلیل عمیق پیکربندی و فایل‌ها — سؤالات قبل از تست تعاملی…",
        )
        try:
            await self._run_planning_phase(agent, attached_file)
        except Exception as exc:  # noqa: BLE001
            await self._publish_phase(
                agent_id,
                "planning",
                f"تحلیل با خطا مواجه شد: {type(exc).__name__}: {exc}",
            )
            # Soft-fail: mark analysis empty so training can still proceed
            await self._store_planning(
                agent_id,
                analysis="تحلیل خودکار در دسترس نبود؛ با نمونه‌ها و دستورالعمل ادامه دهید.",
                questions=[],
                awaiting_answers=False,
            )
        agent = await self.db.get(Agent, agent_id)
        return agent or fresh

    async def _run_planning_phase(self, agent: Agent, attached_file: AgentFile | None) -> bool:
        """Analyze config/files; pause validation if clarifying questions are needed."""
        action_rows = await self._list_actions(agent.id)
        file_previews = await self._gather_file_previews(agent.id, attached_file)
        payload = {
            "agent_name": agent.name,
            "kind": agent.kind.value if hasattr(agent.kind, "value") else str(agent.kind),
            "system_prompt": agent.system_prompt or "",
            "tool_names": list(agent.tool_names or []),
            "capabilities": agent.capabilities or {},
            "runtime_plan": (agent.config_json or {}).get("runtime_plan") or {},
            "workspace_script": (agent.config_json or {}).get("workspace_script") or {},
            "actions": [
                {
                    "slug": a.slug,
                    "name": getattr(a, "name", None),
                    "label": getattr(a, "label", None),
                    "description": getattr(a, "description", None),
                    "tool_chain": list(getattr(a, "tool_chain", None) or []),
                    "input_schema": getattr(a, "input_schema", None) or {},
                }
                for a in action_rows
            ],
            "files": file_previews,
        }
        human = (
            "این پیکربندی ایجنت را قبل از تست خودکار بررسی کن. "
            "فقط سؤالاتی بپرس که پاسخشان واقعاً روی درستی خروجی اثر دارد.\n"
            "تمام متن‌های analysis و questions باید فارسی روان و ساده باشند.\n\n"
            f"{json.dumps(payload, ensure_ascii=False, indent=2, default=str)}"
        )
        llm = build_llm(agent)
        response = await llm.ainvoke(
            [
                SystemMessage(content=self._PLANNING_SYSTEM),
                HumanMessage(content=human),
            ]
        )
        raw = getattr(response, "content", str(response))
        if isinstance(raw, list):
            raw = "".join(
                part.get("text", "") if isinstance(part, dict) else str(part) for part in raw
            )
        parsed = self._parse_planning_json(str(raw))
        analysis = str(parsed.get("analysis") or "").strip()
        questions = [
            q
            for q in (parsed.get("questions") or [])
            if isinstance(q, dict) and str(q.get("text", "")).strip()
        ][:5]
        for i, q in enumerate(questions):
            q.setdefault("id", f"q{i + 1}")
            q["id"] = str(q["id"])
            q["text"] = str(q.get("text") or "").strip()
            q["context"] = str(q.get("context") or "").strip()
            q["options"] = self._normalize_planning_options(q)

        awaiting = bool(questions)
        await self._store_planning(agent.id, analysis, questions, awaiting_answers=awaiting)
        return awaiting

    @staticmethod
    def _normalize_planning_options(q: dict) -> list[str]:
        """Always return 3–5 short Persian single-select chips (Cursor-style).

        LLMs often omit ``options``; never leave the UI with only «سایر».
        """
        raw = q.get("options") or q.get("choices") or q.get("answers") or []
        if isinstance(raw, str):
            raw = [p.strip() for p in re.split(r"[|؛\n/]+", raw) if p.strip()]
        if not isinstance(raw, list):
            raw = []
        opts = [str(o).strip() for o in raw if str(o or "").strip()][:6]
        opts = [o for o in opts if o not in ("سایر", "سایر…", "other", "Other")]
        if len(opts) >= 2:
            return opts[:5]
        text = str(q.get("text") or "")
        ctx = str(q.get("context") or "")
        hay = f"{text} {ctx}"
        # Domain-aware defaults so chips are useful without a second LLM call.
        # Domain-neutral defaults first; HR/holiday pads only for matching questions.
        if any(k in hay for k in ("ستون", "شیت", "خروجی", "ساختار", "عنوان", "csv", "xlsx")):
            return [
                "عین ساختار فایل خروجی نمونه",
                "طبق دستورالعمل (حتی اگر با نمونه فرق کند)",
                "اولویت با خروجی نمونه",
                "ساده‌سازی ستون‌ها در حد امکان",
            ]
        if any(k in hay for k in ("ورودی", "فایل خام", "نمونه", "اکسل", "csv")):
            return [
                "همین فایل ورودی فعلی کافی است",
                "باید دقیقاً مثل خروجی نمونه باشد",
                "هر دو فایل را با هم در نظر بگیر",
                "فقط قوانین متنی مهم است",
            ]
        if any(k in hay for k in ("تعطیل", "تقویم", "time.ir", "مناسبت", "holiday", "کارکرد", "مرخصی", "موظف")):
            return [
                "از جدول تعطیلات time.ir داخل سیستم",
                "فقط از فایل دستورالعمل",
                "از هر دو (time.ir + دستورالعمل)",
                "طبق فایل خروجی نمونه",
            ]
        if opts:
            base = opts[0]
            return [
                base,
                "طبق دستورالعمل",
                "طبق خروجی نمونه",
                "هر دو با اولویت دستورالعمل",
            ]
        return [
            "طبق دستورالعمل",
            "طبق خروجی/ورودی نمونه",
            "هر دو با اولویت دستورالعمل",
            "تصمیم با ایجنت در حد معقول",
        ]

    async def _clear_planning(self, agent_id) -> None:
        async with async_session_maker() as db:
            agent = await db.get(Agent, agent_id)
            if not agent or agent.status == AgentStatus.PAUSED:
                return
            cfg = dict(agent.config_json or {})
            validation = dict(cfg.get("validation") or {})
            validation.pop("planning", None)
            validation["state"] = "running"
            validation["current_phase"] = "starting"
            cfg["validation"] = validation
            agent.config_json = cfg
            await db.commit()

    async def _store_planning(
        self,
        agent_id,
        analysis: str,
        questions: list[dict],
        *,
        awaiting_answers: bool,
    ) -> None:
        async with async_session_maker() as db:
            agent = await db.get(Agent, agent_id)
            if not agent or agent.status == AgentStatus.PAUSED:
                return
            cfg = dict(agent.config_json or {})
            validation = dict(cfg.get("validation") or {})
            validation["state"] = "planning" if awaiting_answers else "running"
            validation["current_phase"] = "planning"
            # Re-normalize options at store time so older LLM payloads and
            # partial JSON never persist questions with options=null.
            safe_questions: list[dict] = []
            for i, q in enumerate(questions):
                if not isinstance(q, dict):
                    continue
                item = dict(q)
                item.setdefault("id", f"q{i + 1}")
                item["id"] = str(item["id"])
                item["text"] = str(item.get("text") or "").strip()
                item["context"] = str(item.get("context") or "").strip()
                item["options"] = AgentValidationService._normalize_planning_options(item)
                if item["text"]:
                    safe_questions.append(item)
            validation["planning"] = {
                "analysis": analysis,
                "questions": safe_questions,
                "awaiting_answers": awaiting_answers,
                "locale": self.PLANNING_LOCALE,
            }
            cfg["validation"] = validation
            agent.config_json = cfg
            from sqlalchemy.orm.attributes import flag_modified

            flag_modified(agent, "config_json")
            await db.commit()

    @staticmethod
    def _parse_planning_json(text: str) -> dict:
        text = (text or "").strip()
        if text.startswith("```"):
            text = re.sub(r"^```(?:json)?\s*", "", text)
            text = re.sub(r"\s*```$", "", text)
        try:
            parsed = json.loads(text)
            if isinstance(parsed, dict):
                return parsed
        except json.JSONDecodeError:
            pass
        match = re.search(r"\{[\s\S]*\}", text)
        if match:
            try:
                parsed = json.loads(match.group(0))
                if isinstance(parsed, dict):
                    return parsed
            except json.JSONDecodeError:
                pass
        return {"analysis": text[:500], "questions": []}

    async def _gather_file_previews(
        self, agent_id, attached_file: AgentFile | None, *, limit: int = 5
    ) -> list[dict]:
        result = await self.db.execute(
            select(AgentFile)
            .where(AgentFile.agent_id == agent_id)
            .order_by(desc(AgentFile.created_at))
            .limit(limit)
        )
        rows = list(result.scalars().all())
        if not rows and attached_file:
            rows = [attached_file]
        previews: list[dict] = []
        for row in rows:
            previews.append(
                {
                    "filename": row.filename,
                    "mime_type": row.mime_type,
                    "storage_path": row.storage_path,
                    "preview": self._extract_file_preview(row),
                }
            )
        return previews

    def _extract_file_preview(self, row: AgentFile, *, max_chars: int = 3000) -> str:
        path = Path(row.storage_path)
        if not path.is_file():
            return ""
        raw = path.read_bytes()
        lower = row.filename.lower()
        mime = row.mime_type or ""

        if lower.endswith((".xlsx", ".xls")) or "spreadsheet" in mime:
            try:
                wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
                lines: list[str] = []
                for sheet in wb.worksheets[:2]:
                    lines.append(f"[sheet: {sheet.title}]")
                    for i, row_vals in enumerate(sheet.iter_rows(values_only=True)):
                        if i >= 12:
                            lines.append("…")
                            break
                        cells = [str(c) if c is not None else "" for c in row_vals]
                        if any(cells):
                            lines.append("\t".join(cells))
                text = "\n".join(lines)
            except Exception:
                text = ""
        else:
            try:
                text = extract_text(raw, mime, row.filename) or ""
            except Exception:
                text = raw.decode("utf-8", errors="ignore") if mime.startswith("text/") else ""

        text = (text or "").strip()
        if len(text) > max_chars:
            return text[: max_chars - 1] + "…"
        return text

    async def _ensure_sample_file(self, agent: Agent) -> AgentFile:
        filename, mime, raw = self._build_sample_file(agent)
        await validate_upload(self.db, agent, filename=filename, mime_type=mime, size_bytes=len(raw))

        base_dir = Path("var/agent_files") / str(agent.id)
        base_dir.mkdir(parents=True, exist_ok=True)
        storage_name = f"{uuid4().hex}_{filename}"
        storage_path = base_dir / storage_name
        storage_path.write_bytes(raw)

        from src.core.agent_file_roles import ROLE_RUNTIME

        row = AgentFile(
            agent_id=agent.id,
            filename=filename,
            mime_type=mime,
            size_bytes=len(raw),
            storage_path=str(storage_path),
            role=ROLE_RUNTIME,
        )
        self.db.add(row)
        await self.db.flush()
        await self.db.refresh(row)
        return row

    def _build_sample_file(self, agent: Agent) -> tuple[str, str, bytes]:
        from src.core.file_policy import resolve_io_policies

        inp, _out = resolve_io_policies(agent)
        exts = [str(e).lower() for e in inp.allowed_extensions]
        tools = {t.lower() for t in (agent.tool_names or [])}
        # Prefer headers from config io_examples / io_schema when present.
        cfg = agent.config_json or {}
        io_ex = cfg.get("io_examples") or {}
        if isinstance(io_ex, dict) and io_ex.get("input_text"):
            return (
                "sample.txt",
                "text/plain",
                str(io_ex["input_text"])[:8000].encode("utf-8", errors="replace"),
            )

        if "run_agent_script" in tools or ".xlsx" in exts or ".xls" in exts:
            return self._xlsx_sample(agent)
        if ".csv" in exts:
            return ("sample.csv", "text/csv", b"id,name,value\n1,alpha,10\n2,beta,20\n")
        if ".txt" in exts or not exts:
            return ("sample.txt", "text/plain", b"Sample validation file for agent smoke test.")
        if ".pdf" in exts:
            return ("sample.txt", "text/plain", b"Fallback sample for validation.")
        return ("sample.txt", "text/plain", b"Sample validation file.")

    def _xlsx_sample(self, agent: Agent | None = None) -> tuple[str, str, bytes]:
        """Domain-neutral smoke workbook (not HR/karkard-shaped)."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        headers = ["id", "name", "qty", "price"]
        # Optional: first sheet headers from io_schema if available
        try:
            schema = ((agent.config_json or {}).get("io_schema") or {}).get("input") if agent else None
            sheets = (schema or {}).get("sheets") or []
            if sheets and sheets[0].get("headers"):
                headers = [str(h) for h in sheets[0]["headers"] if h is not None] or headers
                if sheets[0].get("name"):
                    ws.title = str(sheets[0]["name"])[:31]
        except Exception:  # noqa: BLE001
            pass
        ws.append(headers)
        # two dummy rows matching header width
        ws.append([1, "alpha", 2, 10][: len(headers)] + [""] * max(0, len(headers) - 4))
        ws.append([2, "beta", 3, 20][: len(headers)] + [""] * max(0, len(headers) - 4))
        buff = io.BytesIO()
        wb.save(buff)
        return (
            "sample.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            buff.getvalue(),
        )

    def _smoke_prompt(self, agent: Agent, attached_file: AgentFile | None) -> str:
        tools = {str(t).lower() for t in (agent.tool_names or [])}
        clarification = self._planning_clarification_block(agent)
        if attached_file:
            ctx = json.dumps(
                {
                    "storage_path": attached_file.storage_path,
                    "agent_id": str(agent.id),
                },
                ensure_ascii=False,
            )
            if "run_agent_script" in tools:
                base = (
                    "Automatic validation run. "
                    f"Context for tools (use these exact values when calling tools):\n{ctx}\n"
                    "Process the uploaded spreadsheet with run_agent_script "
                    "and return a short confirmation with the download link."
                )
                return f"{base}\n\n{clarification}".strip() if clarification else base
            base = (
                "This is an automatic validation run. "
                f"Context for tools (use these exact values when calling tools):\n{ctx}\n"
                "Use the already uploaded file and return a short result confirming execution."
            )
            return f"{base}\n\n{clarification}".strip() if clarification else base
        if "resume_screen" in tools:
            base = (
                "Automatic validation run. You MUST call resume_screen with "
                'role="Backend Engineer" and min_score=6, then return a one-line summary '
                "with how many candidates passed."
            )
            return f"{base}\n\n{clarification}".strip() if clarification else base
        base = "This is an automatic validation run. Return a one-line successful response."
        return f"{base}\n\n{clarification}".strip() if clarification else base

    @staticmethod
    def _planning_clarification_block(agent: Agent) -> str:
        planning = ((agent.config_json or {}).get("validation") or {}).get("planning") or {}
        answers = planning.get("answers") or {}
        questions = planning.get("questions") or []
        if not answers or not questions:
            return ""
        qmap = {
            str(q.get("id")): str(q.get("text", ""))
            for q in questions
            if isinstance(q, dict) and q.get("id")
        }
        lines = ["User clarified the following before testing:"]
        for qid, answer in answers.items():
            qtext = qmap.get(str(qid), str(qid))
            lines.append(f"Q: {qtext}")
            lines.append(f"A: {answer}")
        return "\n".join(lines)

    @staticmethod
    def _schema_properties(schema: dict | None) -> dict:
        if not isinstance(schema, dict):
            return {}
        props = schema.get("properties")
        if isinstance(props, dict):
            return props
        skip = frozenset({"properties", "required", "type", "$schema"})
        return {
            k: v
            for k, v in schema.items()
            if k not in skip and isinstance(v, dict)
        }

    def _action_variables(self, action: AgentAction) -> dict:
        props = self._schema_properties(action.input_schema)
        tool_chain = [str(t) for t in (getattr(action, "tool_chain", None) or [])]
        needs_role = "resume_screen" in tool_chain

        out: dict = {}
        for key, meta in props.items():
            if isinstance(meta, dict) and "default" in meta:
                out[key] = meta["default"]
                continue
            t = meta.get("type") if isinstance(meta, dict) else "string"
            if t in ("integer", "number"):
                out[key] = 1
            elif t == "boolean":
                out[key] = True
            else:
                out[key] = "sample"

        if needs_role:
            role = str(out.get("role", "")).strip()
            if not role or role.lower() in {"sample", "python", "test"}:
                out["role"] = "Backend Engineer"
        return out

    async def _notify_owner_result(
        self, agent: Agent, owner: User, failures: list[ValidationFailure]
    ) -> None:
        testing_link = f"/agents/create/testing?slug={agent.slug}"
        fix_link = f"/agents/{agent.slug}/fix"
        if not failures:
            await self.notifications.create(
                user_id=owner.id,
                title=f"ایجنت «{agent.name}» آماده است",
                message="تست خودکار با موفقیت انجام شد. می‌توانید ایجنت را باز کنید.",
                severity=NotificationSeverity.SUCCESS,
                link=testing_link,
                meta={
                    "agent_id": str(agent.id),
                    "agent_slug": agent.slug,
                    "validation_ok": True,
                },
            )
            return

        fixable = [f for f in failures if f.fixable_in_admin]
        if fixable:
            short = "; ".join(f"{f.phase}: {f.message[:100]}" for f in fixable[:2])
            await self.notifications.create(
                user_id=owner.id,
                title=f"خطا در تست ایجنت «{agent.name}»",
                message=f"تنظیمات ایجنت نیاز به اصلاح دارد. {short}",
                severity=NotificationSeverity.ERROR,
                link=fix_link,
                meta={
                    "agent_id": str(agent.id),
                    "agent_slug": agent.slug,
                    "validation_ok": False,
                    "fixable": True,
                },
            )
            return

        await self.notifications.create(
            user_id=owner.id,
            title=f"تست ایجنت «{agent.name}» ناموفق بود",
            message="مشکل موقت یا زیرساختی رخ داد. بعداً دوباره تست کنید یا با پشتیبانی تماس بگیرید.",
            severity=NotificationSeverity.WARNING,
            link=testing_link,
            meta={
                "agent_id": str(agent.id),
                "agent_slug": agent.slug,
                "validation_ok": False,
                "fixable": False,
            },
        )

    async def _notify_admins_if_fixable(self, agent: Agent, failures: list[ValidationFailure]) -> None:
        fixable = [f for f in failures if f.fixable_in_admin]
        if not fixable:
            return

        result = await self.db.execute(select(User).where(User.is_superuser.is_(True), User.is_active.is_(True)))
        admins = list(result.scalars().all())
        if not admins:
            return

        short = "; ".join(f"{f.phase}: {f.message[:120]}" for f in fixable[:3])
        if len(fixable) > 3:
            short += f" (+{len(fixable) - 3} more)"

        for admin in admins:
            await self.notifications.create(
                user_id=admin.id,
                title=f"Agent validation failed: {agent.name}",
                message=f"Fixable configuration issues detected. {short}",
                severity=NotificationSeverity.ERROR,
                link=f"/agents/{agent.slug}/fix",
                meta={
                    "agent_id": str(agent.id),
                    "agent_slug": agent.slug,
                    "fixable_count": len(fixable),
                },
            )

    async def _validate_tool_resolution(self, agent: Agent, owner: User) -> None:
        """Every configured tool must resolve and bind to the ReAct graph.

        Catches bad slugs / unbindable tools as a fixable config issue instead
        of a silent runtime no-op when the agent is later invoked.
        """
        from src.agents_lib.graph_agent import resolve_bound_tools

        names = await self.orchestrator.resolve_tool_names(agent, owner)
        if not names:
            return
        _tools, missing = resolve_bound_tools(names)
        if missing:
            raise ValueError(
                "ابزارهای ثبت‌نشده در پیکربندی: "
                + "، ".join(sorted(missing))
                + " — آن‌ها را از فهرست ابزارها/اقدام‌ها حذف یا اصلاح کنید."
            )

    async def _validate_instruction_compile(self, agent: Agent) -> None:
        result = await self.db.execute(
            select(AgentFile).where(AgentFile.agent_id == agent.id)
        )
        inst_files = [
            row for row in result.scalars().all() if is_instruction_file(row.filename)
        ]
        if not inst_files:
            return

        cfg = dict(agent.config_json or {})
        meta = dict(cfg.get("instruction_prompt") or {})
        status = str(meta.get("status") or "")
        if status not in {"ready", "fallback"}:
            raise ValueError(
                "فایل دستورالعمل پیوست شده اما system prompt کامپایل نشده — "
                "دوباره منتشر کنید یا دستورالعمل را در ویرایش ذخیره کنید."
            )
        if int(meta.get("rule_count") or 0) < 1:
            raise ValueError("استخراج قوانین از فایل دستورالعمل خالی بود.")

        keywords: set[str] = set()
        for row in inst_files:
            path = Path(row.storage_path)
            if not path.is_file():
                continue
            text = extract_text(path.read_bytes(), row.mime_type, row.filename) or ""
            for token in ("پنجشنبه", "جمعه", "اضافه", "کسر", "موظف", "ستون"):
                if token in text:
                    keywords.add(token)
        prompt = (agent.system_prompt or "").lower()
        if keywords:
            missing = [k for k in keywords if k not in prompt]
            if len(missing) == len(keywords):
                raise ValueError(
                    "دستورالعمل کامپایل شده به نظر ناقص است — قوانین کلیدی فایل در system prompt نیست."
                )

    async def _run_with_timeout_retry(self, task_factory):
        last_exc: Exception | None = None
        for _ in range(self.step_retries + 1):
            try:
                return await asyncio.wait_for(task_factory(), timeout=self.step_timeout_s)
            except asyncio.TimeoutError as exc:
                last_exc = exc
                continue
        if last_exc:
            raise last_exc

    # ponytail: outer rounds loop removed — single verify() pass now (see
    # comment in validate_after_create). _stamp_planning_answers is called
    # inline before the single pass; no re-stamp loop. Kept for the import
    # graph and any future per-action stamping need.
    def _stamp_planning_answers(self, agent: Agent, planning: dict) -> None:
        qs = planning.get("questions") or []
        answers = planning.get("answers") or {}
        if not qs or not answers:
            return
        lines = []
        for q in qs:
            if not isinstance(q, dict):
                continue
            qid = str(q.get("id", ""))
            qtext = str(q.get("text", "")).strip()
            ans = str(answers.get(qid, "")).strip()
            if qtext and ans:
                lines.append(f"Q: {qtext}\nA: {ans}")
        if not lines:
            return
        cfg = dict(agent.config_json or {})
        cfg["planning_answers_context"] = "\n".join(lines)
        agent.config_json = cfg
        flag_modified(agent, "config_json")

    def _is_fixable(self, exc: Exception) -> bool:
        if isinstance(exc, asyncio.TimeoutError):
            return False

        if isinstance(exc, HTTPException):
            return 400 <= int(exc.status_code) < 500

        if isinstance(exc, AppError):
            if exc.code in (ErrorCode.LLM_UNAVAILABLE, ErrorCode.SERVICE_UNAVAILABLE):
                return False
            if exc.code == ErrorCode.ORCHESTRATION_FAILED:
                detail_type = ""
                if isinstance(exc.details, dict):
                    detail_type = str(exc.details.get("type") or "")
                non_fixable_types = {
                    "PermissionError",
                    "ConnectTimeout",
                    "TimeoutException",
                    "ReadTimeout",
                    "TimeoutError",
                }
                return detail_type not in non_fixable_types
            return exc.status_code < 500

        name = type(exc).__name__
        if name in {"BadRequestError", "ValidationError", "NotImplementedError", "KeyError"}:
            return True

        msg = str(exc).lower()
        fixable_tokens = (
            "tool",
            "model",
            "file type not allowed",
            "max file",
            "permission denied",
            "chat is disabled",
            "actions are disabled",
            "max agent call depth",
            "deprecated for this model",
        )
        return any(token in msg for token in fixable_tokens)
