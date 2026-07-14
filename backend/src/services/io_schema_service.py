"""Agent-agnostic I/O schema fingerprints for script synthesis and verification."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _cell_dtype(value: Any) -> str:
    if value is None or value == "":
        return "empty"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, (int, float)):
        return "number"
    text = str(value).strip()
    if not text:
        return "empty"
    if "/" in text and text.replace("/", "").replace(" ", "").isdigit():
        return "dateish"
    if ":" in text and all(p.isdigit() for p in text.split(":") if p):
        return "timeish"
    try:
        float(text.replace(",", "").replace("٫", "."))
        return "number"
    except ValueError:
        return "string"


def _trim_row(row: tuple[Any, ...] | list[Any]) -> list[Any]:
    out = list(row)
    while out and (out[-1] is None or out[-1] == ""):
        out.pop()
    return out


def _xlsx_sheet_schema(ws, *, head: int = 8, tail: int = 6) -> dict[str, Any]:
    rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(_trim_row(row))
    # drop fully empty trailing rows for counts
    while rows and all(c is None or c == "" for c in rows[-1]):
        rows.pop()
    headers = [str(c) if c is not None else "" for c in (rows[0] if rows else [])]
    body = rows[1:] if len(rows) > 1 else []
    dtypes: list[str] = []
    if headers:
        for col_i in range(len(headers)):
            samples = [_cell_dtype(r[col_i]) if col_i < len(r) else "empty" for r in body[:20]]
            # majority non-empty
            non_empty = [s for s in samples if s != "empty"]
            dtypes.append(max(set(non_empty), key=non_empty.count) if non_empty else "empty")
    head_rows = rows[:head]
    tail_rows = rows[-tail:] if len(rows) > head else []
    return {
        "name": ws.title,
        "row_count": len(rows),
        "col_count": max((len(r) for r in rows), default=0),
        "headers": headers,
        "dtypes": dtypes,
        "head": head_rows,
        "tail": tail_rows,
    }


def schema_for_path(path: Path | str, *, budget_chars: int = 24000) -> dict[str, Any]:
    """Build a structural fingerprint for an input or output sample file."""
    path = Path(path)
    suffix = path.suffix.lower()
    base: dict[str, Any] = {
        "path": path.name,
        "suffix": suffix,
        "format": "unknown",
        "sheets": [],
    }
    try:
        if suffix == ".xlsx":
            base["format"] = "xlsx"
            try:
                wb = load_workbook(path, data_only=True, read_only=True)
            except Exception:
                wb = load_workbook(path, data_only=True, read_only=True)
            sheets = []
            for sn in wb.sheetnames:
                try:
                    sheets.append(_xlsx_sheet_schema(wb[sn]))
                except Exception as exc:  # noqa: BLE001
                    sheets.append({"name": sn, "error": f"{type(exc).__name__}: {exc}"})
            try:
                wb.close()
            except Exception:  # noqa: BLE001
                pass
            base["sheets"] = sheets
            base["sheet_names"] = [s.get("name") for s in sheets]
            base["sheet_row_counts"] = {
                s.get("name"): s.get("row_count") for s in sheets if s.get("name")
            }
        elif suffix == ".csv":
            base["format"] = "csv"
            with path.open(newline="", encoding="utf-8-sig") as f:
                rows = [_trim_row(r) for r in csv.reader(f)]
            while rows and all(c is None or c == "" for c in rows[-1]):
                rows.pop()
            headers = [str(c) for c in (rows[0] if rows else [])]
            body = rows[1:] if len(rows) > 1 else []
            dtypes = []
            for col_i in range(len(headers)):
                samples = [_cell_dtype(r[col_i]) if col_i < len(r) else "empty" for r in body[:20]]
                non_empty = [s for s in samples if s != "empty"]
                dtypes.append(max(set(non_empty), key=non_empty.count) if non_empty else "empty")
            base["sheets"] = [
                {
                    "name": path.stem,
                    "row_count": len(rows),
                    "col_count": max((len(r) for r in rows), default=0),
                    "headers": headers,
                    "dtypes": dtypes,
                    "head": rows[:12],
                    "tail": rows[-6:] if len(rows) > 12 else [],
                }
            ]
            base["sheet_names"] = [path.stem]
            base["sheet_row_counts"] = {path.stem: len(rows)}
        elif suffix in {".txt", ".md", ".json"}:
            text = path.read_text(encoding="utf-8", errors="replace")
            base["format"] = "text"
            base["char_count"] = len(text)
            base["line_count"] = text.count("\n") + (1 if text else 0)
            base["preview"] = text[:4000]
        elif suffix == ".pdf":
            base["format"] = "pdf"
            base["size_bytes"] = path.stat().st_size
            try:
                from src.services.agent_file_service import AgentFileService  # type: ignore

                # best-effort text extract without full service
                raw = path.read_bytes()
                # lazy: store size only; extractors live elsewhere
                base["preview"] = f"(pdf {len(raw)} bytes)"
            except Exception:  # noqa: BLE001
                base["preview"] = f"(pdf {path.stat().st_size} bytes)"
        else:
            try:
                text = path.read_text(encoding="utf-8", errors="replace")
                base["format"] = "text"
                base["preview"] = text[:2000]
            except Exception:  # noqa: BLE001
                base["size_bytes"] = path.stat().st_size
    except Exception as exc:  # noqa: BLE001
        base["error"] = f"{type(exc).__name__}: {exc}"

    packed = json.dumps(base, ensure_ascii=False, default=str)
    if len(packed) > budget_chars:
        # shrink head/tail first
        for sheet in base.get("sheets") or []:
            if isinstance(sheet, dict):
                sheet["head"] = (sheet.get("head") or [])[:4]
                sheet["tail"] = (sheet.get("tail") or [])[:3]
        packed = json.dumps(base, ensure_ascii=False, default=str)
        base["_truncated"] = len(packed) > budget_chars
        if len(packed) > budget_chars:
            base["head_only"] = True
            for sheet in base.get("sheets") or []:
                if isinstance(sheet, dict):
                    sheet.pop("tail", None)
                    sheet["head"] = (sheet.get("head") or [])[:3]
    return base


def schema_prompt_block(schema: dict[str, Any] | None, *, label: str) -> str:
    if not schema:
        return f"{label}: (missing)"
    return f"{label}:\n" + json.dumps(schema, ensure_ascii=False, default=str)[:20000]


def build_io_schema_pair(
    input_path: Path | None,
    output_path: Path | None,
) -> dict[str, Any]:
    return {
        "input": schema_for_path(input_path) if input_path and Path(input_path).is_file() else None,
        "output": schema_for_path(output_path) if output_path and Path(output_path).is_file() else None,
    }


def persist_io_schema(agent, schema: dict[str, Any]) -> None:
    """Write schema into agent.config_json.io_schema (caller flushes)."""
    from sqlalchemy.orm.attributes import flag_modified

    cfg = dict(agent.config_json or {})
    cfg["io_schema"] = schema
    agent.config_json = cfg
    try:
        flag_modified(agent, "config_json")
    except Exception:  # noqa: BLE001
        pass
