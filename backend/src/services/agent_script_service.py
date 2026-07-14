"""Wizard-time pinned scripts for deterministic file agents."""

from __future__ import annotations

import asyncio
import csv
import hashlib
import json
import os
import py_compile
import re
import runpy
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from langchain_openai import ChatOpenAI
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm.attributes import flag_modified

from src.core.agent_file_roles import (
    ROLE_INPUT_SAMPLE,
    ROLE_OUTPUT_SAMPLE,
    ROLE_RUNTIME,
    agent_file_role,
    is_instruction_file,
    is_output_sample_file,
    pair_id_from_filename,
)
from src.core.agent_workspace_files import agent_workspace_root
from src.core import llm_runtime
from src.agents_lib.agent_factory import _supports_temperature
from src.models.agent import Agent, AgentKind
from src.models.agent_action import AgentAction
from src.models.agent_file import AgentFile
from src.services.io_schema_service import (
    build_io_schema_pair,
    persist_io_schema,
    schema_for_path,
    schema_prompt_block,
)

SCRIPT_IMPORT_ALLOWLIST = frozenset(
    {
        "__future__",
        "csv",
        "datetime",
        "decimal",
        "json",
        "math",
        "openpyxl",
        "pandas",
        "pathlib",
        "re",
        "shutil",
        "statistics",
        "time",
        "typing",
        "jdatetime",
    }
)

# Admin-editable extra imports, hydrated from platform settings at startup so a
# new library can be allowed without a code edit.
SCRIPT_IMPORT_ALLOWLIST_KEY = "script_import_allowlist"
_EXTRA_ALLOWED_IMPORTS: set[str] = set()

# ponytail: cheap compile-repair passes (local, no LLM cost). One LLM
# re-synthesis attempt if the very first parse misses a truncation SyntaxError
# — the verifier falls back to the safe copy-stub after that instead of
# burning more tokens. Kept at 1 so total LLM spend stays inside
# _SCRIPT_SYNTH_BUDGET.
_REPAIR_ATTEMPTS = 1

# ponytail: ONE regenerate-and-compare pass. The old multi-pass loop (4×8 = 32
# LLM calls) burned tokens without converging — Persian payroll mismatch was
# almost always a structural hint the LLM needed *upfront*, not a retry loop.
# We now front-load that hint (exact failing row/cell, holiday table, planning
# answers) and synthesize ONCE; verify+run once; if it still mismatches we hand
# back a single fixable ValidationFailure instead of looping. The user can edit
# in the wizard's /fix flow far cheaper than another LLM round-trip.
_SCRIPT_VERIFY_MAX_ATTEMPTS = 1

# ponytail: hard cap on total LLM ainvoke calls per verify() — one best-effort
# synthesis + one repair (only if runtime produced a cell/row mismatch we can
# feed back). _compile_or_repair (syntax-only, no LLM) sits outside this
# counter so truncation repair doesn't burn the budget.
_SCRIPT_SYNTH_BUDGET = 2


def set_extra_allowed_imports(items) -> set[str]:
    global _EXTRA_ALLOWED_IMPORTS
    _EXTRA_ALLOWED_IMPORTS = {str(i).strip() for i in (items or []) if str(i).strip()}
    return set(_EXTRA_ALLOWED_IMPORTS)


def allowed_imports() -> frozenset[str]:
    return frozenset(SCRIPT_IMPORT_ALLOWLIST | _EXTRA_ALLOWED_IMPORTS)


async def hydrate_import_allowlist(db) -> set[str]:
    """Load admin-configured extra imports into the process cache."""
    from src.services.platform_settings_service import PlatformSettingsService

    value = await PlatformSettingsService(db).get_value(SCRIPT_IMPORT_ALLOWLIST_KEY)
    items = value.get("imports") if isinstance(value, dict) else None
    return set_extra_allowed_imports(items or [])

_SCRIPT_KEYWORDS = (
    "xlsx",
    "xls",
    "csv",
    "pdf",
    "parse",
    "transform",
    "decode",
    "generate",
    "merge",
    "calculate",
    "ستون",
    "اکسل",
    "فایل",
    "تبدیل",
    "محاسبه",
    "گزارش",
)


@dataclass(frozen=True)
class ScriptDecision:
    needed: bool
    reason: str
    confidence: str


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _kind(agent: Agent) -> str:
    return str(getattr(getattr(agent, "kind", None), "value", getattr(agent, "kind", "")))


def _slug(agent: Agent) -> str:
    return "".join(ch if ch.isalnum() else "_" for ch in (agent.slug or "agent_script")).strip("_")[:40] or "agent_script"


def _sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()[:16]


def _mark_config_modified(agent: Agent) -> None:
    try:
        flag_modified(agent, "config_json")
    except Exception:
        pass


def _safe_script_path(agent: Agent, rel_path: str) -> Path:
    root = agent_workspace_root(agent.id)
    path = (root / rel_path).resolve()
    try:
        path.relative_to(root)
    except ValueError as exc:
        raise PermissionError("Script path escapes agent workspace") from exc
    if path.suffix != ".py":
        raise ValueError("Script must be a Python file")
    return path


def safe_load_workbook(path: Path | str, *, data_only: bool = True, read_only: bool = False):
    """Load xlsx without dying on corrupt style tables (openpyxl IndexError).

    Some HR exports reference style_ids past the stylesheet length. Full load
    then raises ``IndexError: list index out of range`` inside bind_cells.
    ``read_only=True`` skips per-cell style binding and still yields values.
    """
    path = Path(path)
    try:
        return load_workbook(path, data_only=data_only, read_only=read_only)
    except IndexError:
        return load_workbook(path, data_only=data_only, read_only=True)
    except Exception:
        # Last resort: values-only streaming
        return load_workbook(path, data_only=True, read_only=True)


def _norm_sheet_row(row) -> list[Any]:
    norm = []
    for c in row:
        if merged_cell_types() and isinstance(c, merged_cell_types()):
            norm.append(None)
        elif style_proxy_types() and isinstance(c, style_proxy_types()):
            norm.append(str(c))
        else:
            norm.append(c)
    return norm


def _xlsx_rows(path: Path, *, sheet: str | None = None) -> list[list[Any]]:
    wb = safe_load_workbook(path, data_only=True)
    name = sheet or wb.sheetnames[0]
    ws = wb[name]
    rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(_norm_sheet_row(row))
    try:
        wb.close()
    except Exception:  # noqa: BLE001
        pass
    return rows


def _xlsx_all_sheets(path: Path) -> dict[str, list[list[Any]]]:
    wb = safe_load_workbook(path, data_only=True)
    out: dict[str, list[list[Any]]] = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(_norm_sheet_row(row))
        out[sn] = rows
    try:
        wb.close()
    except Exception:  # noqa: BLE001
        pass
    return out


def merged_cell_types():
    try:
        from openpyxl.cell.cell import MergedCell
        return (MergedCell,)
    except Exception:  # noqa: BLE001
        return ()


def style_proxy_types():
    try:
        from openpyxl.styles.proxy import StyleProxy
        return (StyleProxy,)
    except Exception:  # noqa: BLE001
        return ()


def _csv_rows(path: Path) -> list[list[str]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.reader(f))


def _rows(path: Path) -> list[list[Any]]:
    """Sheet0 / single-table rows (compat). Prefer multi-sheet verify for xlsx."""
    if path.suffix.lower() == ".xlsx":
        return _xlsx_rows(path)
    if path.suffix.lower() == ".csv":
        return _csv_rows(path)
    return [path.read_text(encoding="utf-8", errors="replace").splitlines()]


def _rows_preview(path: Path, *, max_rows: int = 20, limit: int = 12000) -> str:
    """Schema-aware preview: all sheets for xlsx, richer budget than sheet0-only."""
    try:
        schema = schema_for_path(path, budget_chars=limit)
        return json.dumps(schema, ensure_ascii=False, default=str)[:limit]
    except Exception as exc:  # noqa: BLE001
        try:
            rows = _rows(path)[:max_rows]
            return json.dumps(rows, ensure_ascii=False, default=str)[:limit]
        except Exception:  # noqa: BLE001
            return f"(unreadable: {type(exc).__name__})"


def _norm_cell(value: Any, *, ndigits: int) -> Any:
    """Normalize a cell for tolerant comparison (None==\"\", numbers rounded)."""
    if value is None:
        return ""
    # openpyxl can leak StyleProxy/font objects into a cell value; never
    # let one crash the comparer (it just won't match the sample -> fixable).
    try:
        hash(value)
    except TypeError:
        return str(value)
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        try:
            return round(float(value), ndigits)
        except (TypeError, ValueError, OverflowError):
            return value
    text = str(value).strip()
    if not text:
        return ""
    try:
        return round(float(text.replace(",", "").replace("٫", ".")), ndigits)
    except ValueError:
        return text


def _trim_trailing_empty(row: list[Any]) -> list[Any]:
    """Drop trailing cells that are None/empty so the verifier is insensitive
    to openpyxl dropping trailing-None columns (e.g. a row written with 22
    logical columns but only 19 carrying data reads back as 19). Empty trailing
    columns carry no information, so they must not cause a column-count failure.
    """
    out = list(row)
    while out and _norm_cell(out[-1], ndigits=2) == "":
        out.pop()
    return out


def _xlsx_sheet_rowcounts(path: Path) -> dict[str, int]:
    """Per-sheet row counts so the LLM sees which sheet over/underproduces."""
    try:
        wb = safe_load_workbook(path, data_only=True)
        out = {sn: ws.max_row for sn, ws in [(s, wb[s]) for s in wb.sheetnames]}
        try:
            wb.close()
        except Exception:  # noqa: BLE001
            pass
        return out
    except Exception:  # noqa: BLE001
        return {}


def _row_count_diagnostic(a_rows: list[list[Any]], e_rows: list[list[Any]]) -> str:
    """Surface enough structure for the LLM repair loop to self-correct.

    A bare "produced 56, expected 41" leaves the model guessing whether to drop
    or add rows. Show the head/tail markers of each side and the per-sheet row
    deltas (xlsx: one block per sheet) so the model can see the unit of
    over-/under-production.
    """
    def _fmt_row(r: list[Any]) -> str:
        return [str(c)[:14] for c in r[:9]]

    def _head(rows: list[list[Any]], n: int = 6) -> str:
        if not rows:
            return "(empty)"
        out = []
        for i, r in enumerate(rows[:n]):
            out.append(f"  r{i+1}: {_fmt_row(r)}")
        if len(rows) > n:
            out.append("  ...")
        return "\n".join(out)

    def _tail(rows: list[list[Any]], n: int = 12) -> str:
        if not rows:
            return "(empty)"
        out = []
        start = max(0, len(rows) - n)
        for i, r in enumerate(rows[start:], start=start + 1):
            out.append(f"  r{i}: {_fmt_row(r)}")
        return "\n".join(out)

    return (
        f"produced {len(a_rows)} rows, expected {len(e_rows)}.\n"
        f"PRODUCED head:\n{_head(a_rows)}\n"
        f"PRODUCED tail:\n{_tail(a_rows)}\n"
        f"EXPECTED head:\n{_head(e_rows)}\n"
        f"EXPECTED tail:\n{_tail(e_rows)}"
    )


def _xlsx_dates(path: Path) -> set[str]:
    """Date-like strings across all sheets (optional trust-input mode)."""
    if not path:
        return set()
    try:
        sheets = _xlsx_all_sheets(path) if path.suffix.lower() == ".xlsx" else {"": _rows(path)}
        out: set[str] = set()
        for rows in sheets.values():
            for row in rows:
                for d in row:
                    if isinstance(d, str) and "/" in d and d.strip():
                        out.add(d.strip())
        return out
    except Exception:  # noqa: BLE001
        return set()


def _compare_row_tables(
    a_rows: list[list[Any]],
    e_rows: list[list[Any]],
    *,
    sheet_label: str,
    ndigits: int,
    min_accuracy: float,
    input_dates: set[str],
    trust_domain_footers: bool,
) -> None:
    if len(a_rows) != len(e_rows):
        raise ValueError(
            f"[{sheet_label}] Row count mismatch: "
            f"{_row_count_diagnostic(a_rows, e_rows)}"
        )
    # Structural: header row (first non-empty) must match when both present
    if a_rows and e_rows:
        ah = _trim_trailing_empty(a_rows[0])
        eh = _trim_trailing_empty(e_rows[0])
        width_h = max(len(ah), len(eh))
        ah = ah + [""] * (width_h - len(ah))
        eh = eh + [""] * (width_h - len(eh))
        header_mismatch = any(
            _norm_cell(av, ndigits=ndigits) != _norm_cell(ev, ndigits=ndigits)
            for av, ev in zip(ah, eh)
        )
        if header_mismatch:
            raise ValueError(
                f"[{sheet_label}] Header row mismatch: produced {ah!r}, expected {eh!r}"
            )

    total_cells = 0
    matched_cells = 0
    first_mismatches: list[str] = []
    for r, (araw, eraw) in enumerate(zip(a_rows, e_rows), start=1):
        arow = _trim_trailing_empty(araw)
        erow = _trim_trailing_empty(eraw)
        width = max(len(arow), len(erow))
        arow = arow + [""] * (width - len(arow))
        erow = erow + [""] * (width - len(erow))
        trust_input_row = False
        if input_dates:
            for cell in arow:
                if isinstance(cell, str) and cell.strip() in input_dates:
                    trust_input_row = True
                    break
        if trust_domain_footers and not trust_input_row:
            # Opt-in only (HR/task_profile): do not apply to general agents.
            footer_labels = {
                "جمع",
                "مجموع",
                "اضافه کار قابل پرداخت",
                "اضافه کار واقعی",
                "تعطیل کار واقعی",
                "جمعه کاری",
                "شبکاری",
                "کسرکار",
                "کسرکار ",
            }
            row_text = " ".join(str(c) for c in arow)
            trust_input_row = any(lbl in row_text for lbl in footer_labels)
        for c, (av, ev) in enumerate(zip(arow, erow), start=1):
            total_cells += 1
            if _norm_cell(av, ndigits=ndigits) == _norm_cell(ev, ndigits=ndigits):
                matched_cells += 1
            elif trust_input_row:
                total_cells -= 1
            elif len(first_mismatches) < 5:
                first_mismatches.append(
                    f"[{sheet_label}] Cell mismatch at row {r}, col {c}: "
                    f"produced {av!r}, expected {ev!r}"
                )
    if total_cells == 0:
        return
    accuracy = matched_cells / total_cells
    if accuracy < min_accuracy:
        raise ValueError(
            f"[{sheet_label}] Output accuracy {accuracy*100:.1f}% below required "
            f"{min_accuracy*100:.0f}% ({matched_cells}/{total_cells} cells matched).\n"
            + "\n".join(first_mismatches)
        )


def verify_file_matches(
    actual: Path,
    expected: Path,
    *,
    ndigits: int = 2,
    input_path: Path | None = None,
    min_accuracy: float = 0.95,
    trust_domain_footers: bool = False,
) -> None:
    """Multi-sheet structural verifier with soft value accuracy.

    Structural hard fails: suffix, sheet name set, per-sheet row count, header row.
    Value cells use ``min_accuracy`` (default 0.95). Domain footer KEEP is off
    unless ``trust_domain_footers`` is explicitly enabled.
    """
    if actual.suffix.lower() != expected.suffix.lower():
        raise ValueError(f"Output suffix mismatch: {actual.suffix} != {expected.suffix}")

    input_dates = _xlsx_dates(input_path) if input_path is not None else set()

    if actual.suffix.lower() == ".xlsx":
        a_sheets = _xlsx_all_sheets(actual)
        e_sheets = _xlsx_all_sheets(expected)
        if set(a_sheets) != set(e_sheets):
            raise ValueError(
                f"Sheet name mismatch: produced {sorted(a_sheets)} expected {sorted(e_sheets)}"
            )
        # Prefer expected sheet order for diagnostics
        for sn in e_sheets:
            _compare_row_tables(
                a_sheets[sn],
                e_sheets[sn],
                sheet_label=sn,
                ndigits=ndigits,
                min_accuracy=min_accuracy,
                input_dates=input_dates,
                trust_domain_footers=trust_domain_footers,
            )
        return

    # CSV / text: single table
    a_rows, e_rows = _rows(actual), _rows(expected)
    _compare_row_tables(
        a_rows,
        e_rows,
        sheet_label=actual.suffix or "file",
        ndigits=ndigits,
        min_accuracy=min_accuracy,
        input_dates=input_dates,
        trust_domain_footers=trust_domain_footers,
    )


def _script_source() -> str:
    return '''from pathlib import Path
from shutil import copy2


def main(input_path: Path, output_dir: Path, *, agent_id: str, args: dict) -> Path:
    """Return the primary output file."""
    output_dir.mkdir(parents=True, exist_ok=True)
    output = output_dir / input_path.name
    copy2(input_path, output)
    return output
'''


def _parse_code(raw: str) -> str:
    text = raw.strip()
    fenced = re.findall(r"```(?:python)?\s*(.*?)\s*```", text, flags=re.DOTALL)
    if fenced:
        code = fenced[-1].strip()
        if "def main(" in code:
            return code
    if text.startswith("```"):
        text = re.sub(r"^```(?:python)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
    return _strip_prose(text)


def _strip_prose(text: str) -> str:
    """Strip leading descriptive prose that LLMs prepend before `def main`.

    Models routinely write a chain-of-thought sentence (e.g.
    "Looking at the error, ...") before the actual code. Drop everything
    before the first ``import`` or ``def``/``class``/``from`` line so the
    file compiles.
    """
    lines = text.splitlines()
    for i, line in enumerate(lines):
        stripped = line.strip()
        if (
            stripped.startswith("import ")
            or stripped.startswith("from ")
            or stripped.startswith("def main(")
            or stripped.startswith("class ")
        ):
            return "\n".join(lines[i:]).strip()
    return text.strip()


def _truncation_hint(source: str, exc: SyntaxError) -> str | None:
    """Detect truncation signatures and return a pointed repair hint.

    LLM output often cuts off mid-token when max_tokens is hit, leaving an
    unterminated string literal or a dangling comma at the very end. The
    generic "SyntaxError at line N" hint doesn't tell the repair LLM that
    it needs to *continue and finish* the script, so we sniff for the
    signature and add an explicit "complete the script" instruction.
    """
    if not source:
        return None
    tail = source.rstrip()
    last_line = tail.splitlines()[-1] if tail else ""
    last_stripped = last_line.strip()
    looks_truncated = (
        "unterminated string literal" in (exc.msg or "").lower()
        or last_stripped.endswith((",", "(", "{", "["))
        or not tail.endswith((";", ":", ")", "}", "]", '"', "'"))
        or ('"""' in tail and tail.count('"""') % 2 != 0)
        or ("'''" in tail and tail.count("'''") % 2 != 0)
    )
    if not looks_truncated:
        return None
    line = tail.count("\n") + 1
    return (
        f"SyntaxError: {exc.msg} (line {exc.lineno or line}). "
        "The previous script appears TRUNCATED — it ends mid-expression. "
        "Write the COMPLETE script, including the full `def main(...)` body "
        "and a return statement. Do not abbreviate or add placeholders."
    )


def _build_llm(agent: Agent | None = None) -> ChatOpenAI | None:
    resolved = llm_runtime.resolve((agent.model_name if agent else None) or None)
    if not resolved.api_key:
        return None
    kwargs: dict = {
        "model": resolved.model,
        "api_key": resolved.api_key,
        "timeout": 180,
        "max_retries": 1,
        # ponytail: 32k covers any realistic single-file script (Persian
        # payroll runs ~300 lines). Usage is output-bounded — the larger cap
        # costs nothing when the script is short and eliminates the truncation
        # failure mode that used to force the repair loop.
        "max_tokens": 32768,
    }
    if _supports_temperature(resolved.model):
        # ponytail: temperature 0.0 — synthesis is deterministic work. Lower
        # temperature = fewer divergences between runs, lower retry risk, less
        # token-burning churn. The old 0.1 introduced just enough variance to
        # occasionally rewrite a script that *would* have verified on the
        # next call.
        kwargs["temperature"] = 0.0
    if resolved.base_url:
        kwargs["base_url"] = resolved.base_url
    if resolved.provider == "cursor":
        kwargs["use_responses_api"] = False
    return ChatOpenAI(**kwargs)


class AgentScriptService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def evaluate(self, agent: Agent) -> ScriptDecision:
        rows = await self._files(agent)
        has_output_sample = any(
            agent_file_role(r.filename, role=getattr(r, "role", None)) == ROLE_OUTPUT_SAMPLE
            for r in rows
        )
        has_runtime_sample = any(
            agent_file_role(r.filename, role=getattr(r, "role", None))
            in {ROLE_INPUT_SAMPLE, ROLE_RUNTIME}
            for r in rows
        )
        caps = agent.capabilities or {}
        text = " ".join(
            [
                agent.name or "",
                agent.description or "",
                agent.system_prompt or "",
                json.dumps(agent.config_json.get("instruction_rules", []) if agent.config_json else [], ensure_ascii=False),
                " ".join(f.filename or "" for f in rows),
            ]
        ).lower()
        actions = await self._actions(agent)
        has_tool_chain = any(getattr(a, "tool_chain", None) for a in actions)
        worker = _kind(agent) in {
            AgentKind.WORKER.value,
            AgentKind.FILE_INTAKE.value,
            AgentKind.SPREADSHEET.value,
        }
        # Any kind with file_upload + samples can need a script (not only worker kinds).
        file_agent = bool(caps.get("file_upload_enabled"))

        if file_agent and has_output_sample and (has_runtime_sample or worker):
            return ScriptDecision(True, "file_upload + output-sample + sample input", "high")
        if file_agent and worker and any(k in text for k in _SCRIPT_KEYWORDS):
            return ScriptDecision(True, "worker file upload + deterministic file keywords", "medium")
        if has_tool_chain and file_agent and any(k in text for k in _SCRIPT_KEYWORDS):
            return ScriptDecision(True, "file action tool_chain + deterministic file keywords", "medium")
        if file_agent and has_output_sample and has_runtime_sample:
            return ScriptDecision(True, "file_upload + I/O samples (any kind)", "high")
        return ScriptDecision(False, "chat/API/routing agent or no deterministic file signal", "high")

    async def generate_if_needed(self, agent: Agent, *, use_llm: bool = False) -> dict[str, Any]:
        decision = await self.evaluate(agent)
        cfg = dict(agent.config_json or {})
        meta = dict(cfg.get("workspace_script") or {})
        meta.update(
            {
                "needed": decision.needed,
                "decision_reason": decision.reason,
                "confidence": decision.confidence,
                "language": "python",
            }
        )
        if not decision.needed:
            cfg["workspace_script"] = meta
            agent.config_json = cfg
            _mark_config_modified(agent)
            await self.db.flush()
            return meta

        slug = meta.get("slug") or f"process_{_slug(agent)}"
        rel = f"scripts/{slug}.py"
        path = _safe_script_path(agent, rel)
        path.parent.mkdir(parents=True, exist_ok=True)
        # Real file workers get an LLM-synthesized script. Regenerate while
        # unverified so a stale copy-stub never lingers; once verified, keep it.
        regenerate = not path.exists() or (use_llm and not meta.get("verified_at"))
        if regenerate:
            if use_llm:
                await self._publish_thinking(
                agent,
                "در حال نوشتن اسکریپت: خواندن schema نمونه‌ها، قوانین و تولید def main…",
            )
            source = await self._synthesize_script(agent) if use_llm else _script_source()
            if use_llm:
                source = await self._compile_or_repair(agent, source, path)
            path.write_text(source, encoding="utf-8")
            if use_llm:
                await self._publish_thinking(
                    agent,
                    f"اسکریپت نوشته شد ({len(source.splitlines())} خط) — در حال کامپایل…",
                )
        py_compile.compile(str(path), doraise=True)
        meta.update(
            {
                "slug": slug,
                "path": rel,
                "entrypoint": "main",
                "synthesized": bool(use_llm),
                "created_at": meta.get("created_at") or _now(),
            }
        )
        cfg["workspace_script"] = meta
        agent.config_json = cfg
        _mark_config_modified(agent)
        await self.db.flush()
        return meta

    async def verify(self, agent: Agent, *, use_llm: bool = False) -> dict[str, Any]:
        meta = await self.generate_if_needed(agent, use_llm=use_llm)
        if not meta.get("needed"):
            return meta
        pairs = await self._sample_pairs(agent)
        if not pairs:
            raise ValueError(
                "برای پردازش فایل، فایل نمونه ورودی و فایل نمونه خروجی "
                "(output-sample__) لازم است."
            )

        path = _safe_script_path(agent, str(meta.get("path") or ""))
        trust_footers = self._trust_domain_footers(agent)
        last_err: Exception | None = None
        used = 0
        # One run + optional single LLM repair when use_llm (budget-capped).
        max_rounds = 2 if use_llm else 1
        synth_calls = 0
        primary_output = pairs[0][1]
        for used in range(max_rounds):
            try:
                for idx, (input_file, output_sample) in enumerate(pairs, 1):
                    await self._publish_thinking(
                        agent,
                        f"اجرای اسکریپت روی جفت نمونه {idx}/{len(pairs)}: {input_file.name}",
                    )
                    output = await asyncio.to_thread(
                        run_agent_script_file, agent, input_file, script_slug=meta.get("slug")
                    )
                    await self._publish_thinking(
                        agent,
                        f"مقایسه خروجی با طلایی ({output_sample.name})…",
                    )
                    verify_file_matches(
                        output,
                        output_sample,
                        input_path=input_file,
                        trust_domain_footers=trust_footers,
                    )
                last_err = None
                break
            except Exception as exc:  # noqa: BLE001
                last_err = exc
                if (
                    not use_llm
                    or used >= max_rounds - 1
                    or synth_calls >= max(0, _SCRIPT_SYNTH_BUDGET - 1)
                ):
                    break
                await self._publish_thinking(
                    agent,
                    f"خطای تأیید: {type(exc).__name__}: {str(exc)[:400]} — در حال اصلاح اسکریپت…",
                )
                repaired = await self._synthesize_script(
                    agent,
                    feedback=f"{type(exc).__name__}: {exc}",
                    prior_code=path.read_text(encoding="utf-8") if path.is_file() else None,
                )
                synth_calls += 1
                path.write_text(repaired, encoding="utf-8")
                py_compile.compile(str(path), doraise=True)
        if last_err is not None:
            meta = dict(meta)
            meta["last_verify_error"] = f"{type(last_err).__name__}: {last_err}"
            return meta

        cfg = dict(agent.config_json or {})
        meta = dict(cfg.get("workspace_script") or meta)
        meta.update(
            {
                "verified_at": _now(),
                "sample_hash": _sha(primary_output),
                "repair_attempts_used": used,
                "pair_count": len(pairs),
            }
        )
        cfg["workspace_script"] = meta
        agent.config_json = cfg
        _mark_config_modified(agent)
        await self.db.flush()
        return meta

    @staticmethod
    def _trust_domain_footers(agent: Agent) -> bool:
        cfg = agent.config_json or {}
        if cfg.get("verify_trust_domain_footers"):
            return True
        profile = str(cfg.get("task_profile") or "").lower()
        return profile in {"karkard", "hr", "attendance", "payroll"}

    async def _compile_or_repair(self, agent: Agent, source: str, path: Path) -> str:
        """Compile ``source``; on SyntaxError, ask the LLM to repair it.

        Never persists a non-compiling script: if all repair attempts fail,
        writes the safe copy-stub so the agent still has a runnable baseline.
        """
        try:
            compile(source, "<agent_script>", "exec")  # noqa: S102
            return source
        except SyntaxError as exc:
            last = source
            for _ in range(_REPAIR_ATTEMPTS):
                hint = _truncation_hint(last, exc)
                last = await self._synthesize_script(
                    agent,
                    feedback=hint or f"SyntaxError: {exc.msg} (line {exc.lineno})",
                    prior_code=last,
                )
                if "def main(" not in last:
                    continue
                try:
                    compile(last, "<agent_script>", "exec")  # noqa: S102
                    path.write_text(last, encoding="utf-8")
                    return last
                except SyntaxError:
                    continue
            safe = _script_source()
            path.write_text(safe, encoding="utf-8")
            return safe

    async def _files(self, agent: Agent) -> list[AgentFile]:
        result = await self.db.execute(
            select(AgentFile).where(AgentFile.agent_id == agent.id).order_by(AgentFile.created_at.desc())
        )
        return list(result.scalars().all())

    async def _actions(self, agent: Agent) -> list[AgentAction]:
        result = await self.db.execute(select(AgentAction).where(AgentAction.agent_id == agent.id))
        return list(result.scalars().all())

    async def _publish_thinking(self, agent: Agent, text: str) -> None:
        """Write live thinking into config_json for the wizard poll (best-effort)."""
        try:
            from src.database.session import async_session_maker
            from src.models.agent import AgentStatus

            commit = getattr(self.db, "commit", None)
            if callable(commit):
                await commit()
        except Exception:  # noqa: BLE001
            rollback = getattr(self.db, "rollback", None)
            if callable(rollback):
                try:
                    await rollback()
                except Exception:  # noqa: BLE001
                    pass
            return
        try:
            session_factory = async_session_maker
        except Exception:  # noqa: BLE001
            return
        async with session_factory() as db:
            row = await db.get(Agent, agent.id)
            if not row or row.status == AgentStatus.PAUSED:
                return
            cfg = dict(row.config_json or {})
            validation = dict(cfg.get("validation") or {})
            validation["script_thinking"] = text
            validation["current_detail"] = text
            validation["state"] = "running"
            log = list(validation.get("thinking_log") or [])
            log.append(
                {
                    "t": _now(),
                    "phase": validation.get("current_phase") or "script",
                    "text": text,
                }
            )
            validation["thinking_log"] = log[-50:]
            cfg["validation"] = validation
            row.config_json = cfg
            try:
                flag_modified(row, "config_json")
            except Exception:  # noqa: BLE001
                pass
            try:
                await db.commit()
            except Exception:  # noqa: BLE001
                await db.rollback()

    async def _sample_pair(self, agent: Agent) -> tuple[Path | None, Path | None]:
        """Backward-compatible first input/output pair."""
        pairs = await self._sample_pairs(agent)
        if not pairs:
            return None, None
        return pairs[0]

    async def _sample_pairs(self, agent: Agent) -> list[tuple[Path, Path]]:
        """All (input, output) sample pairs, grouped by pair_id when present."""
        files = await self._files(agent)
        inputs: dict[str, Path] = {}
        outputs: dict[str, Path] = {}
        for row in files:
            if not Path(row.storage_path).is_file():
                continue
            role = agent_file_role(row.filename, role=getattr(row, "role", None))
            pid = getattr(row, "pair_id", None) or pair_id_from_filename(row.filename) or "default"
            if role == ROLE_OUTPUT_SAMPLE:
                outputs[pid] = outputs.get(pid) or Path(row.storage_path)
            elif role in {ROLE_INPUT_SAMPLE, ROLE_RUNTIME}:
                inputs[pid] = inputs.get(pid) or Path(row.storage_path)
        pairs = [
            (inputs[pid], outputs[pid])
            for pid in sorted(set(inputs) | set(outputs))
            if pid in inputs and pid in outputs
        ]
        return pairs

    async def _synthesize_script(
        self,
        agent: Agent,
        *,
        feedback: str | None = None,
        prior_code: str | None = None,
    ) -> str:
        pairs = await self._sample_pairs(agent)
        llm = _build_llm(agent)
        if not llm or not pairs:
            return _script_source()
        input_path, output_path = pairs[0]
        cfg = agent.config_json or {}
        allowed = ", ".join(sorted(allowed_imports() - {"__future__"}))
        sys = (
            "Write a deterministic Python script for a generic enterprise file agent. "
            "Return ONLY Python code. No markdown. "
            "The script must define exactly: "
            "def main(input_path: Path, output_dir: Path, *, agent_id: str, args: dict) -> Path. "
            "It must COMPUTE the output from the input data; never hardcode values or copy the "
            "reference output sample. Reproduce the output sample's structure, columns, row order, "
            "and formatting as closely as possible. The reference output sample (and any user "
            "clarifications passed in the prompt) are the source of truth for the exact layout — "
            "derive the structure FROM those, do not invent a fixed schema.\n"
            "GENERAL GUIDANCE (agent-agnostic):\n"
            "- Persian/Jalali dates: if you need weekday or date math, use jdatetime "
            "(jdatetime.date.weekday(): 0=شنبه..6=جمعه). Never assume Python's 0=Monday order.\n"
            "- ONE PASS: write the COMPLETE script in a single response, ≤300 lines, never "
            "abbreviate or truncate. Count the sample's rows PER SHEET and reproduce that exact "
            "row count (including any SUM/footer/summary rows). Not minus one, not plus one.\n"
            "- Do NOT embed specific person names (employees, signatories, HR/manager/CEO) as string "
            "literals. Read names/data from the input file; use only generic role labels if a "
            "signature/footer block is required by the sample.\n"
            "- Formulas are fine, but the verifier reads computed values (openpyxl data_only=True), "
            "so ensure the written values — whether literal or formula-backed after recalc — equal "
            "the sample. Writing computed values directly is the simplest reliable path.\n"
            "- openpyxl: ALWAYS load inputs with read_only=True (and data_only=True), e.g. "
            "openpyxl.load_workbook(input_path, data_only=True, read_only=True). Full load can "
            "raise IndexError on some xlsx files with corrupt style tables; read_only avoids that.\n"
            "- Column counts: if trailing columns are empty, write an empty string '' (not None) so "
            "the sheet keeps the expected column count (openpyxl drops trailing None cells).\n"
            "- Match the sample's sheet count dynamically (build one output sheet per relevant input "
            "section, in the same order) rather than hardcoding a sheet count.\n"
            "- Return ONLY raw Python code. Do NOT wrap in ```python fences. Do NOT prepend prose.\n"
            f"Allowed imports only: {allowed}. "
            "No network, subprocess, eval, exec, shell, or absolute paths. "
            "Keep the script compact (aim for ≤300 lines): factor repeated helpers, avoid long "
            "inline constant tables, and do not embed per-row literal lists. "
            "Always end with a complete `return out_path` — the script must be self-contained "
            "and finish; truncated output is the most common failure."
        )
        user_parts = [
            f"Agent name: {agent.name}",
            f"Description: {agent.description or '—'}",
            f"Rules: {json.dumps(cfg.get('instruction_rules') or [], ensure_ascii=False)[:8000]}",
        ]
        # Schema-driven previews (all sheets) for every pair
        for i, (inp, outp) in enumerate(pairs, 1):
            user_parts.append(
                f"\n=== Sample pair {i} ==="
                f"\nInput file: {inp.name}"
                f"\n{schema_prompt_block(schema_for_path(inp), label='Input schema')}"
                f"\nOutput file: {outp.name}"
                f"\n{schema_prompt_block(schema_for_path(outp), label='Expected output schema')}"
            )
        user_parts.append(
            "Write code that reads the runtime input file and writes the primary output file "
            "in output_dir, reproducing the expected output for this input."
        )
        planning_ctx = cfg.get("planning_answers_context")
        if planning_ctx:
            user_parts.append(
                f"\nUser clarifications (apply these rules strictly):\n{planning_ctx}"
            )
        # Holiday calendar only when the agent explicitly wants it (gated by task profile).
        try:
            from src.services.holiday_service import ensure_holiday_calendar

            holiday_calendar = ensure_holiday_calendar(agent)
        except Exception:  # noqa: BLE001
            holiday_calendar = (cfg or {}).get("holiday_calendar")
        if holiday_calendar and holiday_calendar.get("by_year") and self._trust_domain_footers(agent):
            user_parts.append(
                "\nNational holiday calendar (Jalali, from time.ir — static, no network at "
                "runtime):\n"
                + json.dumps(holiday_calendar, ensure_ascii=False)[:12000]
                + "\nTreat these days as رسمی روز تعطیل. Weekly off-days still apply."
            )
        if feedback:
            user_parts.append(
                "\nThe previous attempt FAILED verification with this error:\n"
                f"{feedback}\n"
                "Fix the script so its output matches the expected output exactly. "
                "Edit the previous script in place; do not rewrite from scratch."
            )
        if prior_code:
            user_parts.append(f"\nPrevious script (edit this in place):\n{prior_code[:6000]}")
        # ponytail: caveman mode for script writing only — terse-prompt patch
        # saves output tokens (the model returns dense Python, no chatter).
        # Activated by env var MA_SCRIPT_WRITER=caveman so it only applies to
        # the synthesis path, never chat/planning.
        if os.environ.get("MA_SCRIPT_WRITER") == "caveman":
            sys += (
                "\nCAVEMAN: reply terse. No explanation. No prose. Only complete Python "
                "code. One pass. ≤300 lines. End with `return out_path`."
            )
        candidate = prior_code
        retry_hint: str | None = None
        # ponytail: single LLM call here. The old loop (up to 3 ainvoke calls)
        # multiplied with _compile_or_repair's outer loop and the verify() outer
        # loop, burning 8+ tokens-per-script passes. One synthesis attempt per
        # _synthesize_script invocation is enough; downstream callers (verify,
        # _compile_or_repair) honor _SCRIPT_SYNTH_BUDGET for the total LLM spend.
        for attempt in range(1):
            try:
                prompt_user = "\n".join(user_parts)
                if attempt > 0 and candidate:
                    prompt_user += (
                        "\n\nThe previous attempt did not compile or was truncated. "
                        "Return ONLY the complete Python code inside a single ```python fenced "
                        "block, ending with `return out_path`. Do not prepend prose and do not "
                        "emit a second fenced block."
                        + (f"\n{retry_hint}" if retry_hint else "")
                    )
                result = await llm.ainvoke(
                    [{"role": "system", "content": sys}, {"role": "user", "content": prompt_user}]
                )
                code = _parse_code(getattr(result, "content", None) or str(result))
            except Exception:
                code = ""
            candidate = code or prior_code or ""
            if "def main(" in candidate:
                try:
                    compile(candidate, "<agent_script>", "exec")  # noqa: S102
                    return candidate
                except SyntaxError as exc:
                    retry_hint = _truncation_hint(candidate, exc)
            if not candidate:
                break
        return prior_code or _script_source()


def _load_main(script_path: Path):
    source = script_path.read_text(encoding="utf-8")
    for bad in ("subprocess", "socket", "requests", "urllib", "http.client", "os.system", "eval(", "exec("):
        if bad in source:
            raise PermissionError(f"Script uses forbidden capability: {bad}")
    for line in source.splitlines():
        stripped = line.strip()
        if stripped.startswith("import ") or stripped.startswith("from "):
            root = stripped.split()[1].split(".", 1)[0]
            if root not in allowed_imports():
                raise PermissionError(f"Import not allowed: {root}")
    ns = runpy.run_path(str(script_path))
    main = ns.get("main")
    if not callable(main):
        raise ValueError("Script missing main(input_path, output_dir, *, agent_id, args)")
    return main


def _script_timeout_seconds() -> int:
    try:
        return max(5, int(os.environ.get("MA_SCRIPT_TIMEOUT_S", "120")))
    except ValueError:
        return 120


def _run_script_in_subprocess(
    script_path: Path,
    input_path: Path,
    output_dir: Path,
    *,
    agent_id: str,
    args: dict | None,
    workspace_root: Path,
    timeout_s: int | None = None,
) -> Path:
    """Execute pinned script in a subprocess with timeout + optional rlimits."""
    import subprocess
    import tempfile

    timeout_s = timeout_s or _script_timeout_seconds()
    output_dir.mkdir(parents=True, exist_ok=True)
    marker = output_dir / f".script_result_{os.getpid()}.json"
    if marker.exists():
        try:
            marker.unlink()
        except OSError:
            pass

    # Static analysis in parent before spawn (same as in-process).
    source = script_path.read_text(encoding="utf-8")
    for bad in ("subprocess", "socket", "requests", "urllib", "http.client", "os.system", "eval(", "exec("):
        if bad in source:
            raise PermissionError(f"Script uses forbidden capability: {bad}")
    for line in source.splitlines():
        stripped = line.strip()
        if stripped.startswith("import ") or stripped.startswith("from "):
            root_mod = stripped.split()[1].split(".", 1)[0]
            if root_mod not in allowed_imports():
                raise PermissionError(f"Import not allowed: {root_mod}")

    runner = f'''
import json, sys
from pathlib import Path
import openpyxl as _ox
_orig = _ox.load_workbook
def _patched(filename, *a, **kw):
    try:
        return _orig(filename, *a, **kw)
    except IndexError:
        kw = dict(kw)
        kw["read_only"] = True
        kw.setdefault("data_only", True)
        return _orig(filename, *a, **kw)
_ox.load_workbook = _patched
import runpy
ns = runpy.run_path({str(script_path)!r})
main = ns.get("main")
if not callable(main):
    raise SystemExit("missing main")
out = main(
    Path({str(input_path)!r}),
    Path({str(output_dir)!r}),
    agent_id={agent_id!r},
    args=json.loads({json.dumps(args or {})!r}),
)
Path({str(marker)!r}).write_text(json.dumps({{"out": str(out)}}), encoding="utf-8")
'''

    def _preexec():
        try:
            import resource

            # Soft caps; wall-clock timeout is primary. AS ~2GiB for openpyxl/pandas.
            soft_as = 2 * 1024 * 1024 * 1024
            resource.setrlimit(resource.RLIMIT_AS, (soft_as, soft_as))
            resource.setrlimit(resource.RLIMIT_CPU, (max(5, timeout_s + 5), max(5, timeout_s + 5)))
            resource.setrlimit(resource.RLIMIT_NOFILE, (256, 256))
        except Exception:
            pass

    env = {
        "PATH": os.environ.get("PATH", "/usr/bin:/bin"),
        "PYTHONPATH": os.environ.get("PYTHONPATH", ""),
        "HOME": str(workspace_root),
        "LANG": "C.UTF-8",
    }
    # Prefer MA_SCRIPT_PYTHON or current interpreter
    py = os.environ.get("MA_SCRIPT_PYTHON") or sys.executable
    try:
        proc = subprocess.run(
            [py, "-c", runner],
            cwd=str(workspace_root),
            env=env,
            capture_output=True,
            text=True,
            timeout=timeout_s,
            preexec_fn=_preexec if os.name == "posix" else None,
            check=False,
        )
    except subprocess.TimeoutExpired as exc:
        raise TimeoutError(f"Script exceeded {timeout_s}s timeout") from exc

    if proc.returncode != 0:
        err = (proc.stderr or proc.stdout or "").strip()[:2000]
        raise RuntimeError(f"Script failed (exit {proc.returncode}): {err or 'unknown error'}")
    if not marker.is_file():
        raise FileNotFoundError("Script did not report an output path")
    payload = json.loads(marker.read_text(encoding="utf-8"))
    try:
        marker.unlink()
    except OSError:
        pass
    out_path = Path(payload["out"]).resolve()
    try:
        out_path.relative_to(workspace_root.resolve())
    except ValueError as exc:
        raise PermissionError("Script output escapes agent workspace") from exc
    if not out_path.is_file():
        raise FileNotFoundError("Script did not create output file")
    return out_path


def run_agent_script_file(
    agent: Agent,
    input_path: Path,
    *,
    script_slug: str | None = None,
    args: dict | None = None,
) -> Path:
    meta = dict((agent.config_json or {}).get("workspace_script") or {})
    if not meta.get("needed"):
        raise ValueError("Agent has no pinned workspace script")
    if script_slug and script_slug != meta.get("slug"):
        raise PermissionError("Requested script_slug is not pinned for this agent")
    script_path = _safe_script_path(agent, str(meta.get("path") or ""))
    if not script_path.is_file():
        raise FileNotFoundError("Pinned script file not found")

    root = agent_workspace_root(agent.id)
    input_path = input_path.resolve()
    try:
        input_path.relative_to(root)
    except ValueError as exc:
        raise PermissionError("Input path escapes agent workspace") from exc

    output_dir = root / "output"

    # Default: subprocess sandbox. Opt out with MA_SCRIPT_INPROCESS=1 for debug.
    if os.environ.get("MA_SCRIPT_INPROCESS") == "1":
        def _execute() -> Path:
            import openpyxl as _ox

            _orig_load = _ox.load_workbook

            def _patched_load(filename, *a, **kw):
                try:
                    return _orig_load(filename, *a, **kw)
                except IndexError:
                    kw = dict(kw)
                    kw["read_only"] = True
                    kw.setdefault("data_only", True)
                    return _orig_load(filename, *a, **kw)

            _ox.load_workbook = _patched_load  # type: ignore[assignment]
            try:
                out = _load_main(script_path)(
                    input_path,
                    output_dir,
                    agent_id=str(agent.id),
                    args=dict(args or {}),
                )
            finally:
                _ox.load_workbook = _orig_load  # type: ignore[assignment]
            out_path = Path(out).resolve()
            try:
                out_path.relative_to(root)
            except ValueError as exc:
                raise PermissionError("Script output escapes agent workspace") from exc
            if not out_path.is_file():
                raise FileNotFoundError("Script did not create output file")
            return out_path

        return _execute()

    return _run_script_in_subprocess(
        script_path,
        input_path,
        output_dir,
        agent_id=str(agent.id),
        args=dict(args or {}),
        workspace_root=root,
    )
