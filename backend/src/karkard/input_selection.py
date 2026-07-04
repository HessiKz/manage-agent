"""Pick the raw کارکرد upload among agent files."""

from __future__ import annotations

from pathlib import Path
from typing import Protocol

from src.core.agent_file_roles import is_instruction_file, is_output_sample_file
from src.karkard.names import is_processed_karkard_filename
from src.karkard.processor import _find_header_row, _is_already_processed_sheet


class _FileRow(Protocol):
    filename: str | None
    storage_path: str


def _is_processed_karkard_name(name: str) -> bool:
    return is_processed_karkard_filename(name)


def is_runtime_karkard_candidate(filename: str | None, storage_path: str | None = None) -> bool:
    """True when this attachment could be raw attendance input (not sample/instruction/output)."""
    name = filename or ""
    path_name = Path(storage_path or "").name
    if is_instruction_file(name) or is_output_sample_file(name):
        return False
    if is_instruction_file(path_name) or is_output_sample_file(path_name):
        return False
    if not name.lower().endswith((".xlsx", ".xls")):
        return False
    if _is_processed_karkard_name(name) or _is_processed_karkard_name(path_name):
        return False
    return True


def workbook_looks_like_raw_karkard(path: Path) -> bool:
    """Peek headers — raw exports have timeclock columns; samples/outputs do not."""
    if not path.is_file():
        return False
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            if _is_already_processed_sheet(ws):
                return False
            header_row = _find_header_row(ws)
            if not header_row:
                return False
            headers = {
                str(c.value or "").replace("\n", " ").strip() for c in ws[header_row]
            }
            if "اضافه کار کل" in headers:
                return True
            if "کارکرد" in headers and "تاریخ" in headers:
                return True
            # ponytail: some HR exports omit «اضافه کار کل» label but keep punch columns
            if "اولین ورود" in headers and "آخرین خروج" in headers and "تاریخ" in headers:
                return True
            return False
        finally:
            wb.close()
    except Exception:
        return False


def list_agent_file_rows(agent_id: str) -> list[_FileRow]:
    """Build file rows from on-disk agent uploads (newest first)."""
    from src.core.runtime_file_selection import list_agent_file_rows as _list

    return _list(agent_id)  # type: ignore[return-value]


def pick_runtime_karkard_file(files: list[_FileRow]) -> _FileRow | None:
    """Choose raw کارکرد input only — newest upload wins; never fall back to stale files."""
    candidates = [
        f
        for f in files
        if is_runtime_karkard_candidate(f.filename, f.storage_path)
    ]
    if not candidates:
        return None
    # ponytail: files are newest-first; re-processing an older upload caused repeat outputs
    row = candidates[0]
    path = Path(row.storage_path or "")
    if path.is_file() and workbook_looks_like_raw_karkard(path):
        return row
    return None
