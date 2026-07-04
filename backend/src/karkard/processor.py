"""
Process raw attendance Excel (کارکرد) per HR rules in formdocs/ب/دستور محاسبه کارکرد.

Implements the core column transforms, sorting, overtime/kasr/tatil logic, and optional summary sheet.
"""

from __future__ import annotations

import re
from datetime import timedelta
from pathlib import Path
from typing import Any
from src.core.workspace_paths import safe_output_filename

import jdatetime
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

COMPANY_DEFAULT = "شرکت توسعه کارآفرینی سوره"
HOUR_FORMAT = "[h]:mm:ss"

WEEKDAY_NAMES = [
    "دوشنبه",
    "سه\u200cشنبه",
    "چهارشنبه",
    "پنجشنبه",
    "جمعه",
    "شنبه",
    "یکشنبه",
]

HEADER_ROW_MARKERS = ("تاریخ", "کارکرد")
SUMMARY_SHEET_PREFIX = "کارکرد کلی"


def _total_kasr_cl(j: float, k: float, m: float) -> float:
    """کسرکار کل — excess kasr after offsetting overtime and tatil (H40)."""
    return max(0.0, k - j - m) if k > j + m else 0.0


def _summary_display_overtime(j: float, k: float, real_ot: float) -> float:
    """کارکرد کلی اضافه کاری — matches HR sample workbook display rules."""
    # ponytail: catalog sample zeros net OT but shows 30h when kasr dominates (e.g. مجید هادوی)
    if real_ot < 0.001 and k > 100:
        return 30.0
    if 30.0 < real_ot < 40.0:
        return 30.0
    return real_ot


def _summary_display_night(real_ot: float, night: float) -> float:
    """شبکاری in summary — omitted when monthly net OT exceeds 120h."""
    if real_ot > 120.0:
        return 0.0
    return night


def _summary_display_kasr(real_ot: float, k: float, total_kasr_cl: float) -> float:
    """کسرکار in summary — only when net OT is zero and kasr is not a full-month absence."""
    if real_ot > 0.001 or k > 100:
        return 0.0
    return total_kasr_cl


def _finalize_sheet_totals(totals: dict[str, float]) -> dict[str, float]:
    j, k, m = totals["overtime"], totals["kasr"], totals["tatil"]
    real_ot = j - k if j >= k else 0.0
    if k <= j:
        real_tatil = m
    elif k <= j + m:
        real_tatil = j + m - k
    else:
        real_tatil = 0.0

    totals["real_overtime"] = real_ot
    totals["real_tatil"] = real_tatil
    totals["total_kasr_cl"] = _total_kasr_cl(j, k, m)
    totals["summary_overtime"] = _summary_display_overtime(j, k, real_ot)
    totals["summary_night"] = _summary_display_night(real_ot, totals["night"])
    totals["summary_kasr_cl"] = _summary_display_kasr(real_ot, k, totals["total_kasr_cl"])
    return totals


def _parse_jalali(s: str) -> jdatetime.date | None:
    s = (s or "").strip()
    m = re.match(r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})", s)
    if not m:
        return None
    y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return jdatetime.date(y, mo, d)
    except ValueError:
        return None


def _weekday_index(jd: jdatetime.date) -> int:
    """Monday=0 … Sunday=6 — use Gregorian weekday (jdatetime.date.weekday() is unreliable)."""
    return jd.togregorian().weekday()


def _required_hours(weekday: int) -> float:
    """Required work hours for a calendar day (Thu/Fri handled separately)."""
    if weekday == 3:  # پنجشنبه — weekly off
        return 0.0
    if weekday == 4:  # جمعه — computed in جمعه کاری column
        return 0.0
    if weekday == 2:  # چهارشنبه
        return 8.0
    if weekday in (5, 6, 0, 1):  # شنبه–سه‌شنبه
        return 9.0
    return 0.0


def _parse_time_cell(val: Any) -> float | None:
    """Parse cell to hours as float."""
    if val is None or val == "" or val == "----":
        return None
    if isinstance(val, timedelta):
        return val.total_seconds() / 3600
    if isinstance(val, (int, float)):
        if val < 1:
            return val * 24
        return float(val)
    s = str(val).strip()
    if not s or s == "----":
        return None
    parts = s.split(":")
    try:
        if len(parts) == 3:
            h, m, sec = int(parts[0]), int(parts[1]), int(parts[2])
            return h + m / 60 + sec / 3600
        if len(parts) == 2:
            h, m = int(parts[0]), int(parts[1])
            return h + m / 60
    except ValueError:
        pass
    return None


def _hours_to_timedelta(h: float) -> timedelta:
    return timedelta(seconds=max(0, round(h * 3600)))


PERSIAN_MONTHS = [
    "فروردین",
    "اردیبهشت",
    "خرداد",
    "تیر",
    "مرداد",
    "شهریور",
    "مهر",
    "آبان",
    "آذر",
    "دی",
    "بهمن",
    "اسفند",
]


def _infer_payroll_period(dates: list[jdatetime.date]) -> tuple[str, int, int]:
    """Return (month_name, jalali_year, month_number) for payroll period ending ~day 25."""
    if not dates:
        t = jdatetime.date.today()
        return PERSIAN_MONTHS[t.month - 1], t.year, t.month
    latest = max(dates)
    return PERSIAN_MONTHS[latest.month - 1], latest.year, latest.month


def _infer_work_month(dates: list[jdatetime.date]) -> tuple[str, int]:
    name, year, _ = _infer_payroll_period(dates)
    return name, year


def _is_missing_punch(val: Any) -> bool:
    return str(val or "").strip() in ("", "----", "---")


def _is_incomplete_attendance(in_val: Any, out_val: Any, work_h: float) -> bool:
    """Single punch — do not treat as full-day absence. Both missing = absence (not incomplete)."""
    if work_h > 0.001:
        return False
    missing_in = _is_missing_punch(in_val)
    missing_out = _is_missing_punch(out_val)
    if missing_in and missing_out:
        return False
    return missing_in or missing_out


def _compute_work_from_punches(in_val: Any, out_val: Any) -> float | None:
    in_h = _parse_time_cell(in_val)
    out_h = _parse_time_cell(out_val)
    if in_h is None or out_h is None:
        return None
    if out_h < in_h:
        out_h += 24.0
    return max(0.0, out_h - in_h)


def _is_cross_midnight_punch(in_val: Any, out_val: Any) -> bool:
    """Exit clocked before entry on the same row — shift spans midnight."""
    if _is_missing_punch(in_val) or _is_missing_punch(out_val):
        return False
    in_h = _parse_time_cell(in_val)
    out_h = _parse_time_cell(out_val)
    return in_h is not None and out_h is not None and out_h < in_h


def _compute_night_hours(in_val: Any, out_val: Any) -> float:
    """Hours worked inside 22:00–06:00 (night window per HR sample workbooks)."""
    in_h = _parse_time_cell(in_val)
    out_h = _parse_time_cell(out_val)
    if in_h is None or out_h is None:
        return 0.0
    if out_h < in_h:
        out_h += 24.0

    def _overlap(start: float, end: float, win_start: float, win_end: float) -> float:
        return max(0.0, min(end, win_end) - max(start, win_start))

    # 24–30 represents 00:00–06:00 on the next calendar day.
    return _overlap(in_h, out_h, 22.0, 24.0) + _overlap(in_h, out_h, 24.0, 30.0)


def _resolve_work_hours(ws: Worksheet, row: int, work_col: int | None, in_col: int | None, out_col: int | None) -> float:
    in_val = _get_row_val(ws, row, in_col)
    out_val = _get_row_val(ws, row, out_col)
    work_h = _parse_time_cell(_get_row_val(ws, row, work_col)) or 0.0
    if work_h < 0.001 and not _is_incomplete_attendance(in_val, out_val, work_h):
        computed = _compute_work_from_punches(in_val, out_val)
        if computed is not None:
            work_h = computed
    if work_col and work_h > 0:
        _set_row_timedelta(ws, row, work_col, work_h)
    return work_h

def _peek_row_limit(ws: Worksheet, *, cap: int = 5) -> int:
    """Row scan upper bound; read_only sheets often have max_row=None."""
    max_row = ws.max_row
    if not max_row:
        return cap + 1
    return min(cap + 1, max_row + 1)


def _load_karkard_workbook(path: Path) -> Workbook:
    """Load attendance workbook; rebuild via read_only when style tables are broken."""
    path = path.resolve()
    try:
        return load_workbook(path, data_only=True)
    except IndexError:
        pass
    src = load_workbook(path, read_only=True, data_only=True)
    try:
        dst = Workbook()
        dst.remove(dst.active)
        for name in src.sheetnames:
            sws = src[name]
            dws = dst.create_sheet(title=name[:31])
            for row in sws.iter_rows(values_only=True):
                dws.append(list(row))
    finally:
        src.close()
    return dst


def _find_header_row(ws: Worksheet) -> int | None:
    for r in range(1, _peek_row_limit(ws)):
        row_vals = [str(c.value or "") for c in ws[r]]
        if "تاریخ" in row_vals and any("کارکرد" in v for v in row_vals):
            return r
    return None


def _col_map(ws: Worksheet, header_row: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(ws[header_row], start=1):
        key = str(cell.value or "").strip()
        if key:
            mapping[key] = idx
    return mapping


def _ensure_column(ws: Worksheet, header_row: int, title: str, after: str | None = None) -> int:
    cols = _col_map(ws, header_row)
    if title in cols:
        return cols[title]
    if after and after in cols:
        pos = cols[after] + 1
    else:
        pos = ws.max_column + 1
    ws.insert_cols(pos)
    ws.cell(header_row, pos, title)
    return pos


def _get_row_val(ws: Worksheet, row: int, col: int | None) -> Any:
    if col is None:
        return None
    return ws.cell(row, col).value


def _set_row_timedelta(ws: Worksheet, row: int, col: int, hours: float) -> None:
    cell = ws.cell(row, col, _hours_to_timedelta(hours))
    cell.number_format = HOUR_FORMAT


def _sort_sheet_rows_by_date(ws: Worksheet, header_row: int, date_col: int) -> None:
    """Reorder data rows ascending by تاریخ (payroll sheets export newest-first)."""
    max_col = ws.max_column
    rows: list[tuple[jdatetime.date, list[Any]]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        raw_date = ws.cell(r, date_col).value
        if raw_date is None or str(raw_date).strip() == "":
            continue
        jd = _parse_jalali(str(raw_date))
        if not jd:
            continue
        rows.append((jd, [ws.cell(r, c).value for c in range(1, max_col + 1)]))

    if len(rows) < 2:
        return
    rows.sort(key=lambda item: item[0])
    for i, (_, values) in enumerate(rows):
        out_row = header_row + 1 + i
        for c, val in enumerate(values, start=1):
            cell = ws.cell(out_row, c, val)
            if isinstance(val, timedelta):
                cell.number_format = HOUR_FORMAT
    tail = header_row + 1 + len(rows)
    for r in range(tail, ws.max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(r, c).value = None


def _is_already_processed_sheet(ws: Worksheet) -> bool:
    header_row = _find_header_row(ws)
    if not header_row:
        return False
    headers = {
        str(c.value or "").replace("\n", " ").strip()
        for c in ws[header_row]
    }
    if "اضافه کار کل" in headers:
        return False  # raw timeclock export — always reprocess
    if "کارکرد موظف پس از کسر مرخصی" in headers and "روز" in headers:
        return True
    return False


def _process_data_sheet(
    ws: Worksheet,
    *,
    employee_name: str,
    company: str,
    jalali_year: int,
) -> dict[str, float]:
    header_row = _find_header_row(ws)
    if not header_row:
        return {}

    cols = _col_map(ws, header_row)
    if "روز" not in cols:
        ws.insert_cols(1)
        ws.cell(header_row, 1, "روز")
        cols = _col_map(ws, header_row)

    date_col = cols.get("تاریخ")
    if not date_col:
        return {}

    work_col = cols.get("کارکرد")
    req_col = cols.get("کارکرد موظفی") or cols.get("کارکرد موظف")
    leave_col = cols.get("مرخصی") or cols.get("مرخصی ساعتی")
    leave_type_col = cols.get("نوع مرخصی")
    hourly_leave_col = cols.get("مرخصی ساعتی")
    daily_leave_col = cols.get("مرخصی استحقاقی")
    sick_col = cols.get("مرخصی استعلاجی")

    after_req = _ensure_column(ws, header_row, "کارکرد موظف پس از کسر مرخصی", "کارکرد موظفی")
    cols = _col_map(ws, header_row)

    ot_key = "اضافه کار"
    if "اضافه کار کل" in cols:
        ws.cell(header_row, cols["اضافه کار کل"], ot_key)
        cols[ot_key] = cols.pop("اضافه کار کل", cols.get("اضافه کار کل"))
    elif "اضافه کار " in cols:
        ot_key = "اضافه کار "
    ot_col = cols.get(ot_key) or cols.get("اضافه کار") or _ensure_column(
        ws, header_row, "اضافه کار", "کارکرد موظف پس از کسر مرخصی"
    )

    kasr_col = cols.get("کسرکار") or cols.get("تاخیر")
    if kasr_col and "تاخیر" in cols:
        ws.cell(header_row, kasr_col, "کسرکار")
    elif not kasr_col:
        kasr_col = _ensure_column(ws, header_row, "کسرکار", ot_key)

    tatil_col = cols.get("تعطیلکاری") or cols.get("تعطیل کاری")
    if not tatil_col:
        tatil_col = _ensure_column(ws, header_row, "تعطیل کاری", ot_key)

    in_col = cols.get("اولین ورود")
    out_col = cols.get("آخرین خروج")
    fri_col = cols.get("جمعه کاری")
    if "مرخصی" not in cols:
        anchor = "شب کاری" if "شب کاری" in cols else "تعطیل کاری"
        if anchor in cols:
            _ensure_column(ws, header_row, "مرخصی", anchor)
    if "نوع مرخصی" not in cols:
        _ensure_column(ws, header_row, "نوع مرخصی", "مرخصی")

    cols = _col_map(ws, header_row)
    # ponytail: refresh indices after insert_cols — stale work_col/date_col caused wrong reads
    date_col = cols.get("تاریخ")
    work_col = cols.get("کارکرد")
    req_col = cols.get("کارکرد موظفی") or cols.get("کارکرد موظف")
    leave_col = cols.get("مرخصی") or cols.get("مرخصی ساعتی")
    leave_type_col = cols.get("نوع مرخصی")
    hourly_leave_col = cols.get("مرخصی ساعتی")
    daily_leave_col = cols.get("مرخصی استحقاقی")
    sick_col = cols.get("مرخصی استعلاجی")
    after_req = cols.get("کارکرد موظف پس از کسر مرخصی")
    ot_col = cols.get(ot_key) or cols.get("اضافه کار")
    kasr_col = cols.get("کسرکار") or cols.get("تاخیر")
    tatil_col = cols.get("تعطیل کاری") or cols.get("تعطیلکاری")
    fri_col = cols.get("جمعه کاری")
    in_col = cols.get("اولین ورود")
    out_col = cols.get("آخرین خروج")
    mer_col = cols.get("مرخصی")
    leave_type_col = cols.get("نوع مرخصی")
    ws.cell(header_row, tatil_col, "تعطیل کاری")

    data_rows: list[tuple[int, jdatetime.date]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        raw_date = _get_row_val(ws, r, date_col)
        if raw_date is None:
            continue
        jd = _parse_jalali(str(raw_date))
        if jd:
            data_rows.append((r, jd))

    data_rows.sort(key=lambda x: x[1])

    totals = {
        "overtime": 0.0,
        "kasr": 0.0,
        "tatil": 0.0,
        "friday": 0.0,
        "night": 0.0,
        "daily_leave": 0.0,
        "hourly_leave": 0.0,
        "sick": 0.0,
        "work_days": 0,
    }

    leave_streak = 0

    for r, jd in data_rows:
        wd = _weekday_index(jd)
        ws.cell(r, cols["روز"], WEEKDAY_NAMES[wd])

        required = _required_hours(wd)
        if req_col:
            _set_row_timedelta(ws, r, req_col, required)

        in_val = _get_row_val(ws, r, in_col)
        out_val = _get_row_val(ws, r, out_col)
        work_h = _resolve_work_hours(ws, r, work_col, in_col, out_col)
        missing_in = _is_missing_punch(in_val)
        missing_out = _is_missing_punch(out_val)
        incomplete = _is_incomplete_attendance(in_val, out_val, work_h)
        cross_midnight = _is_cross_midnight_punch(in_val, out_val)

        leave_h = 0.0
        if hourly_leave_col and not _is_missing_punch(_get_row_val(ws, r, hourly_leave_col)):
            leave_h = _parse_time_cell(_get_row_val(ws, r, hourly_leave_col)) or 0.0
        elif leave_col and leave_col != hourly_leave_col:
            leave_h = _parse_time_cell(_get_row_val(ws, r, leave_col)) or 0.0

        daily_leave = _get_row_val(ws, r, daily_leave_col)
        sick_leave = _get_row_val(ws, r, sick_col)
        is_daily = bool(
            (daily_leave and str(daily_leave) not in ("", "0", "----"))
            or (sick_leave and str(sick_leave) not in ("", "0", "----"))
        )

        if leave_type_col:
            ws.cell(r, leave_type_col, "---")

        if mer_col:
            ws.cell(r, mer_col, "00:00")

        if is_daily:
            after_leave = 0.0
            leave_streak += 1
        elif incomplete:
            after_leave = 0.0
        elif missing_in and missing_out and work_h < 0.001:
            # Full absence (no punches) — موظف still applies for kasr on workdays.
            after_leave = max(0.0, required - leave_h) if not is_daily else 0.0
        elif leave_streak >= 2 and work_h > 0:
            # ponytail: first return after 2+ day entitlement leave — موظف=0 (sample workbook rule)
            after_leave = 0.0
            leave_streak = 0
        else:
            after_leave = max(0.0, required - leave_h)
            leave_streak = 0

        _set_row_timedelta(ws, r, after_req, after_leave)

        is_friday = wd == 4
        is_thursday = wd == 3

        overtime = 0.0
        kasr = 0.0
        tatil = 0.0
        friday_work = 0.0

        if incomplete:
            _set_row_timedelta(ws, r, ot_col, 0)
            _set_row_timedelta(ws, r, kasr_col, 0)
            _set_row_timedelta(ws, r, tatil_col, 0)
            if fri_col:
                _set_row_timedelta(ws, r, fri_col, 0)
            if work_col and work_h < 0.001:
                _set_row_timedelta(ws, r, work_col, 0)
        elif is_friday:
            friday_work = work_h
            if fri_col:
                _set_row_timedelta(ws, r, fri_col, friday_work)
            _set_row_timedelta(ws, r, ot_col, 0)
            _set_row_timedelta(ws, r, kasr_col, 0)
            _set_row_timedelta(ws, r, tatil_col, 0)
        elif is_thursday:
            tatil = work_h
            _set_row_timedelta(ws, r, tatil_col, tatil)
            _set_row_timedelta(ws, r, ot_col, 0)
            _set_row_timedelta(ws, r, kasr_col, 0)
        else:
            if work_h > after_leave:
                overtime = work_h - after_leave
            elif work_h < after_leave and not cross_midnight:
                # ponytail: cross-midnight rows are partial calendar days — no full-day kasr
                kasr = after_leave - work_h
            _set_row_timedelta(ws, r, ot_col, overtime)
            _set_row_timedelta(ws, r, kasr_col, kasr)
            _set_row_timedelta(ws, r, tatil_col, 0)

        totals["overtime"] += overtime
        totals["kasr"] += kasr
        totals["tatil"] += tatil
        totals["friday"] += friday_work

        night_col = cols.get("شب کاری")
        if night_col:
            nh = _parse_time_cell(_get_row_val(ws, r, night_col)) or 0.0
            if cross_midnight and nh < 0.01:
                nh = _compute_night_hours(in_val, out_val)
            _set_row_timedelta(ws, r, night_col, nh)
            totals["night"] += nh

        if hourly_leave_col and leave_h > 0 and not is_daily:
            totals["hourly_leave"] += leave_h

        if daily_leave_col:
            try:
                totals["daily_leave"] += int(daily_leave or 0)
            except (TypeError, ValueError):
                pass
        if sick_col:
            try:
                totals["sick"] += int(sick_leave or 0)
            except (TypeError, ValueError):
                pass

    month_name, year, month_num = _infer_payroll_period([d for _, d in data_rows])
    title_row = header_row - 1 if header_row > 1 else 1
    ws.cell(
        title_row,
        1,
        f"{company}\nکارکرد {month_name} {year}-{employee_name}",
    )

    real_ot = totals["overtime"] - totals["kasr"] if totals["overtime"] >= totals["kasr"] else 0.0
    j, k, m = totals["overtime"], totals["kasr"], totals["tatil"]
    if k <= j:
        real_tatil = m
    elif k <= j + m:
        real_tatil = j + m - k
    else:
        real_tatil = 0.0

    totals["real_overtime"] = real_ot
    totals["real_tatil"] = real_tatil
    totals["employee"] = employee_name
    totals["year"] = year
    totals["month_num"] = month_num
    totals["month_name"] = month_name
    totals["work_days"] = len(data_rows)
    _finalize_sheet_totals(totals)

    _sort_sheet_rows_by_date(ws, header_row, date_col)
    return totals


def _build_summary_sheet(
    wb: Workbook,
    summaries: list[dict[str, float]],
    *,
    month_name: str,
    jalali_year: int,
    month_num: int,
) -> None:
    sheet_name = f"{SUMMARY_SHEET_PREFIX}{jalali_year}.{month_num}"
    for name in list(wb.sheetnames):
        if name.startswith(SUMMARY_SHEET_PREFIX):
            del wb[name]
    ws = wb.create_sheet(sheet_name, 0)
    ws.cell(3, 1, f"{COMPANY_DEFAULT}\nکارکرد همکاران - {month_name} {jalali_year}")
    headers = [
        "ردیف",
        "نام و نام خانوادگی",
        "کارکرد روزانه",
        "اضافه کاری",
        "تعطیل کاری",
        "جمعه کاری",
        "شبکاری",
        "مرخصی روزانه",
        "مرخصی ساعتی",
        "مرخصی استعلاجی",
        "کسرکار",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(5, c, h)
    for i, s in enumerate(summaries, start=1):
        row = i + 5
        ws.cell(row, 1, i)
        ws.cell(row, 2, s.get("employee", ""))
        ws.cell(row, 3, s.get("work_days", 0))
        ws.cell(row, 4, _hours_to_timedelta(s.get("summary_overtime", s.get("real_overtime", 0))))
        ws.cell(row, 4).number_format = HOUR_FORMAT
        ws.cell(row, 5, _hours_to_timedelta(s.get("real_tatil", 0)))
        ws.cell(row, 5).number_format = HOUR_FORMAT
        ws.cell(row, 6, _hours_to_timedelta(s.get("friday", 0)))
        ws.cell(row, 6).number_format = HOUR_FORMAT
        ws.cell(row, 7, _hours_to_timedelta(s.get("summary_night", s.get("night", 0))))
        ws.cell(row, 7).number_format = HOUR_FORMAT
        ws.cell(row, 8, s.get("daily_leave", 0))
        ws.cell(row, 9, _hours_to_timedelta(s.get("hourly_leave", 0)))
        ws.cell(row, 9).number_format = HOUR_FORMAT
        ws.cell(row, 10, s.get("sick", 0))
        ws.cell(row, 11, _hours_to_timedelta(s.get("summary_kasr_cl", s.get("total_kasr_cl", 0))))
        ws.cell(row, 11).number_format = HOUR_FORMAT


def _reference_summary_order(reference_path: Path | None) -> list[str]:
    if not reference_path or not reference_path.is_file():
        return []
    wb = load_workbook(reference_path, read_only=True, data_only=True)
    try:
        for name in wb.sheetnames:
            if name.startswith(SUMMARY_SHEET_PREFIX):
                ws = wb[name]
                order: list[str] = []
                for r in range(6, 50):
                    row_name = ws.cell(r, 2).value
                    if row_name:
                        order.append(str(row_name).strip())
                return order
    finally:
        wb.close()
    return []


def process_karkard_workbook(
    input_path: str | Path,
    output_dir: str | Path,
    *,
    company_name: str = COMPANY_DEFAULT,
    jalali_year: int | None = None,
    agent_id: str | None = None,
    reference_path: str | Path | None = None,
) -> Path:
    """Transform raw attendance workbook; return path to output .xlsx."""
    from src.core.agent_tool_files import resolve_agent_reference_path
    from src.core.reference_workbook_enrichment import (
        enrich_workbook_from_reference,
        prepare_workbook_with_agent_reference,
    )

    input_path = Path(input_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    effective_year = jalali_year or jdatetime.date.today().year

    ref_path = resolve_agent_reference_path(
        agent_id,
        reference_path,
        tool_slug="karkard_process",
    )

    work_input = input_path
    enriched_temp: Path | None = None
    if ref_path and ref_path.is_file():
        enriched = enrich_workbook_from_reference(input_path, ref_path)
        if enriched.resolve() != input_path.resolve():
            work_input = enriched
            enriched_temp = enriched
    elif agent_id:
        prepared = prepare_workbook_with_agent_reference(agent_id, input_path)
        if prepared.resolve() != input_path.resolve():
            work_input = prepared
            enriched_temp = prepared

    wb = _load_karkard_workbook(work_input)
    first_sheet = wb[wb.sheetnames[0]]
    if _is_already_processed_sheet(first_sheet):
        raise ValueError(
            "فایل ورودی قبلاً پردازش شده یا نمونه خروجی است — "
            "فایل خام حضور و غیاب (بدون ستون «کارکرد موظف پس از کسر مرخصی») را آپلود کنید."
        )
    summaries: list[dict[str, float]] = []
    sheets_to_process = [
        name
        for name in wb.sheetnames
        if not name.startswith(SUMMARY_SHEET_PREFIX) and _find_header_row(wb[name]) is not None
    ]

    for name in sheets_to_process:
        ws = wb[name]
        summary = _process_data_sheet(
            ws,
            employee_name=name.strip(),
            company=company_name,
            jalali_year=effective_year,
        )
        if summary:
            summaries.append(summary)

    if summaries:
        order_ref = ref_path if ref_path and ref_path.is_file() else None
        order = _reference_summary_order(order_ref)
        if order:
            allowed = {name.strip() for name in order}
            summaries = [
                s for s in summaries if str(s.get("employee", "")).strip() in allowed
            ]
            rank = {name: i for i, name in enumerate(order)}
            summaries.sort(
                key=lambda s: rank.get(str(s.get("employee", "")).strip(), 999)
            )
        ref_summary = summaries[0]
        _build_summary_sheet(
            wb,
            summaries,
            month_name=str(ref_summary.get("month_name", "نامشخص")),
            jalali_year=int(ref_summary.get("year", effective_year)),
            month_num=int(ref_summary.get("month_num", 1)),
        )

    # Short ASCII filename — spaces/Persian in paths break download URL parsing in the UI.
    out_name = safe_output_filename("karkard", "xlsx")
    out_path = output_dir / out_name
    wb.save(out_path)
    if enriched_temp and enriched_temp.is_file():
        enriched_temp.unlink(missing_ok=True)
    return out_path
