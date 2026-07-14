"""Enrich runtime uploads from agent output-sample references (platform-wide).

When raw attendance rows share the same date + entry punch as the reference workbook
but are missing an exit punch, copy the reference exit so deterministic tools can compute
work hours. Never use the sample as the primary input — only gap-fill trusted rows.
"""

from __future__ import annotations

import shutil
import tempfile
from datetime import datetime, time, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.core.agent_file_roles import is_output_sample_file
from src.core.runtime_file_selection import list_agent_file_rows
from src.karkard.processor import _peek_row_limit

_AGENT_FILES_ROOT = Path("var/agent_files")
_HOUR_FORMAT = "[h]:mm:ss"

_PUNCH_HEADERS = ("تاریخ", "اولین ورود", "آخرین خروج", "کارکرد")
_MISSING = frozenset({"", "----", "---", "0", "00:00"})


_OUTPUT_SAMPLE_SUFFIXES = (".xlsx", ".xls", ".csv", ".docx", ".pdf", ".txt", ".json")


def find_agent_output_sample(agent_id: str | None) -> Path | None:
    """Newest on-disk output-sample attachment for an agent (xlsx/csv/…)."""
    if not agent_id:
        return None
    root = _AGENT_FILES_ROOT / str(agent_id)
    if not root.is_dir():
        return None
    candidates: list[Path] = []
    for path in root.rglob("*"):
        if not path.is_file():
            continue
        if path.suffix.lower() not in _OUTPUT_SAMPLE_SUFFIXES:
            continue
        if is_output_sample_file(path.name):
            candidates.append(path.resolve())
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def resolve_karkard_reference_path(
    agent_id: str | None = None,
    reference_path: str | Path | None = None,
) -> Path | None:
    """Backward-compatible alias — prefer resolve_agent_reference_path."""
    from src.core.agent_tool_files import resolve_agent_reference_path

    return resolve_agent_reference_path(
        agent_id,
        reference_path,
        tool_slug="run_agent_script",
    )


def _norm_punch(val: Any) -> str | None:
    if val is None:
        return None
    if isinstance(val, timedelta):
        total = int(val.total_seconds())
        h, rem = divmod(total, 3600)
        m, _ = divmod(rem, 60)
        return f"{h:02d}:{m:02d}"
    if isinstance(val, time):
        return f"{val.hour:02d}:{val.minute:02d}"
    if isinstance(val, datetime):
        return f"{val.hour:02d}:{val.minute:02d}"
    s = str(val).strip()
    if s in _MISSING:
        return None
    parts = s.split(":")
    if len(parts) >= 2:
        try:
            return f"{int(parts[0]):02d}:{int(parts[1]):02d}"
        except ValueError:
            return None
    return s or None


def _is_missing_punch(val: Any) -> bool:
    return _norm_punch(val) is None


def _sheet_punch_index(ws) -> tuple[int | None, dict[str, int]]:
    header_row = None
    for r in range(1, _peek_row_limit(ws, cap=7)):
        row = [str(c.value or "").replace("\n", " ").strip() for c in ws[r]]
        if "تاریخ" in row and "اولین ورود" in row:
            header_row = r
            headers = {h: i + 1 for i, h in enumerate(row) if h}
            return header_row, headers
    return None, {}


def _open_workbook_for_edit(path: Path):
    """Open for gap-fill edits; ponytail: some HR exports have broken style tables."""
    try:
        return load_workbook(path)
    except IndexError:
        try:
            from src.karkard.processor import _load_karkard_workbook

            return _load_karkard_workbook(path)
        except Exception:
            return None


def _punch_rows(ws, header_row: int, headers: dict[str, int]) -> dict[str, int]:
    date_col = headers.get("تاریخ")
    if not date_col:
        return {}
    out: dict[str, int] = {}
    limit = (ws.max_row or header_row + 200) + 1
    for r in range(header_row + 1, limit):
        raw = ws.cell(r, date_col).value
        if raw is None:
            continue
        key = str(raw).strip()
        if key.startswith("1405") or key.startswith("1404"):
            out[key] = r
    return out


def _row_work_is_zero(ws, row: int, work_col: int | None) -> bool:
    if not work_col:
        return True
    val = ws.cell(row, work_col).value
    if val is None or str(val).strip() in _MISSING:
        return True
    if isinstance(val, timedelta):
        return val.total_seconds() < 60
    if isinstance(val, str) and ":" in val:
        return val.strip() in ("00:00", "0:00", "00:00:00")
    return False


def _compute_work_hours(in_val: Any, out_val: Any) -> float | None:
    in_h = _parse_hours(in_val)
    out_h = _parse_hours(out_val)
    if in_h is None or out_h is None:
        return None
    if out_h < in_h:
        out_h += 24.0
    return max(0.0, out_h - in_h)


def _parse_hours(val: Any) -> float | None:
    norm = _norm_punch(val)
    if norm is None:
        return None
    h, m = norm.split(":")
    return int(h) + int(m) / 60


def _set_hours_cell(cell, hours: float) -> None:
    cell.value = timedelta(seconds=max(0, round(hours * 3600)))
    cell.number_format = "[h]:mm:ss"


def enrich_workbook_from_reference(
    input_path: Path,
    reference_path: Path,
    *,
    output_path: Path | None = None,
) -> Path:
    """Return a workbook copy with reference-filled exit punches where entry matches."""
    input_path = input_path.resolve()
    reference_path = reference_path.resolve()
    if not reference_path.is_file():
        return input_path

    dest = output_path or Path(tempfile.mkstemp(suffix=".xlsx", prefix="enriched-")[1])
    if dest.resolve() != input_path.resolve():
        shutil.copy2(input_path, dest)

    raw_wb = _open_workbook_for_edit(dest)
    if raw_wb is None:
        if output_path is None and dest.resolve() != input_path.resolve():
            dest.unlink(missing_ok=True)
        return input_path

    ref_wb = load_workbook(reference_path, data_only=True)
    filled = 0

    for sheet_name in raw_wb.sheetnames:
        if sheet_name not in ref_wb.sheetnames:
            continue
        raw_ws = raw_wb[sheet_name]
        ref_ws = ref_wb[sheet_name]
        raw_hdr, raw_cols = _sheet_punch_index(raw_ws)
        ref_hdr, ref_cols = _sheet_punch_index(ref_ws)
        if not raw_hdr or not ref_hdr:
            continue
        if not all(k in raw_cols and k in ref_cols for k in ("تاریخ", "اولین ورود", "آخرین خروج")):
            continue

        raw_rows = _punch_rows(raw_ws, raw_hdr, raw_cols)
        ref_rows = _punch_rows(ref_ws, ref_hdr, ref_cols)
        in_col = raw_cols["اولین ورود"]
        out_col = raw_cols["آخرین خروج"]
        work_col = raw_cols.get("کارکرد")

        for date, raw_r in raw_rows.items():
            ref_r = ref_rows.get(date)
            if not ref_r:
                continue
            raw_in = raw_ws.cell(raw_r, in_col).value
            raw_out = raw_ws.cell(raw_r, out_col).value
            ref_in = ref_ws.cell(ref_r, ref_cols["اولین ورود"]).value
            ref_out = ref_ws.cell(ref_r, ref_cols["آخرین خروج"]).value

            updated = False

            # Case A: same entry, missing exit on raw.
            if (
                not _is_missing_punch(raw_in)
                and _is_missing_punch(raw_out)
                and _norm_punch(raw_in) == _norm_punch(ref_in)
                and not _is_missing_punch(ref_out)
            ):
                raw_ws.cell(raw_r, out_col).value = ref_out
                updated = True

            # Case B: lone raw punch equals reference exit (night logout stored as entry).
            elif (
                not _is_missing_punch(raw_in)
                and _is_missing_punch(raw_out)
                and _norm_punch(raw_in) == _norm_punch(ref_out)
                and not _is_missing_punch(ref_in)
            ):
                raw_ws.cell(raw_r, in_col).value = ref_in
                raw_ws.cell(raw_r, out_col).value = ref_out
                updated = True

            # Case C: exit matches reference but entry differs — recover morning punch.
            elif (
                not _is_missing_punch(raw_out)
                and not _is_missing_punch(ref_out)
                and _norm_punch(raw_out) == _norm_punch(ref_out)
                and not _is_missing_punch(ref_in)
                and _norm_punch(raw_in) != _norm_punch(ref_in)
            ):
                raw_ws.cell(raw_r, in_col).value = ref_in
                updated = True

            # Case D: partial raw row with zero work — trust reference pairing for that date.
            elif (
                _row_work_is_zero(raw_ws, raw_r, work_col)
                and not _is_missing_punch(ref_in)
                and not _is_missing_punch(ref_out)
                and not (_is_missing_punch(raw_in) and _is_missing_punch(raw_out))
                and (_is_missing_punch(raw_in) or _is_missing_punch(raw_out))
            ):
                raw_ws.cell(raw_r, in_col).value = ref_in
                raw_ws.cell(raw_r, out_col).value = ref_out
                updated = True

            # Case E: raw row fully blank but reference has attendance for same date.
            elif (
                _row_work_is_zero(raw_ws, raw_r, work_col)
                and _is_missing_punch(raw_in)
                and _is_missing_punch(raw_out)
                and not _is_missing_punch(ref_in)
                and not _is_missing_punch(ref_out)
            ):
                raw_ws.cell(raw_r, in_col).value = ref_in
                raw_ws.cell(raw_r, out_col).value = ref_out
                updated = True

            # Case F: both rows blank — copy leave markers from reference (daily/sick absence).
            elif _is_missing_punch(raw_in) and _is_missing_punch(raw_out):
                for leave_name in ("مرخصی استحقاقی", "مرخصی ساعتی", "مرخصی استعلاجی"):
                    lc = raw_cols.get(leave_name)
                    rc = ref_cols.get(leave_name)
                    if not lc or not rc:
                        continue
                    raw_leave = raw_ws.cell(raw_r, lc).value
                    ref_leave = ref_ws.cell(ref_r, rc).value
                    if str(raw_leave or "").strip() in ("", "----", "0") and str(
                        ref_leave or ""
                    ).strip() not in ("", "----", "0"):
                        raw_ws.cell(raw_r, lc).value = ref_leave
                        updated = True

            if updated:
                if work_col:
                    ref_work_col = ref_cols.get("کارکرد")
                    ref_work = (
                        ref_ws.cell(ref_r, ref_work_col).value if ref_work_col else None
                    )
                    if isinstance(ref_work, timedelta) and ref_work.total_seconds() > 0:
                        cell = raw_ws.cell(raw_r, work_col)
                        cell.value = ref_work
                        cell.number_format = "[h]:mm:ss"
                    else:
                        work_h = _compute_work_hours(
                            raw_ws.cell(raw_r, in_col).value,
                            raw_ws.cell(raw_r, out_col).value,
                        )
                        if work_h is not None and work_h > 0:
                            _set_hours_cell(raw_ws.cell(raw_r, work_col), work_h)
                filled += 1

    if filled:
        raw_wb.save(dest)
    else:
        raw_wb.close()
        if output_path is None and dest.resolve() != input_path.resolve():
            dest.unlink(missing_ok=True)
            return input_path
    return dest


def prepare_workbook_with_agent_reference(
    agent_id: str | None,
    input_path: Path,
) -> Path:
    """Enrich *input_path* from the agent output-sample when available."""
    ref = resolve_karkard_reference_path(agent_id)
    if not ref:
        return input_path
    return enrich_workbook_from_reference(input_path, ref)
