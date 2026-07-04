#!/usr/bin/env python3
"""Process formdocs raw کارکرد and assert output matches expected reference."""

from __future__ import annotations

import sys
from datetime import timedelta, time
from pathlib import Path

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[2]
BACKEND = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND))

from src.karkard.processor import _finalize_sheet_totals, process_karkard_workbook  # noqa: E402

RAW = REPO / "formdocs" / "ب" / "کارکرد توسعه کارآفرینی-2.1405.xlsx"
EXPECTED = REPO / "formdocs" / "کارکرد_توسعه_کارآفرینی_1405.2.xlsx"


def _norm(v):
    if v is None:
        return None
    if isinstance(v, timedelta):
        return round(v.total_seconds() / 3600, 4)
    if isinstance(v, time):
        return f"{v.hour:02d}:{v.minute:02d}"
    s = str(v).strip().replace("\n", " ")
    if s in ("", "None"):
        return None
    if s == "---":
        return "---"
    if s in ("00:00", "0:00:00"):
        return 0.0
    if ":" in s:
        p = s.split(":")
        try:
            return round(int(p[0]) + int(p[1]) / 60 + (int(p[2]) / 3600 if len(p) > 2 else 0), 4)
        except ValueError:
            return s
    if isinstance(v, (int, float)):
        return v
    return s


def _hours_from_cell(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, timedelta):
        return val.total_seconds() / 3600
    n = _norm(val)
    return float(n) if isinstance(n, float) else 0.0


def _sheet_headers(ws):
    for r in range(1, 8):
        row = [str(c.value or "").replace("\n", " ").strip() for c in ws[r]]
        if "تاریخ" in row:
            return r, {h: i + 1 for i, h in enumerate(row) if h}
    return None, {}


def _footer_totals_from_days(ws, header_row: int, cols: dict[str, int]) -> dict[str, float]:
    """HR footer formulas (H36/H37/…) from day rows — not stale summary cache."""
    date_col = cols.get("تاریخ")
    if not date_col:
        return {}

    def col(name: str) -> int | None:
        for k, v in cols.items():
            if k.replace("\n", " ") == name:
                return v
        return None

    ot_c = col("اضافه کار")
    kasr_c = col("کسرکار")
    tatil_c = col("تعطیل کاری") or col("تعطیلکاری")
    fri_c = col("جمعه کاری")
    night_c = col("شب کاری")
    daily_c = col("مرخصی استحقاقی")
    hourly_c = col("مرخصی ساعتی")
    sick_c = col("مرخصی استعلاجی")

    j = k = m = l = n = hourly = 0.0
    daily = sick = 0
    work_days = 0
    for r in range(header_row + 1, ws.max_row + 1):
        d = str(ws.cell(r, date_col).value or "").strip()
        if not d.startswith("1405"):
            continue
        work_days += 1
        if ot_c:
            j += _hours_from_cell(ws.cell(r, ot_c).value)
        if kasr_c:
            k += _hours_from_cell(ws.cell(r, kasr_c).value)
        if tatil_c:
            m += _hours_from_cell(ws.cell(r, tatil_c).value)
        if fri_c:
            l += _hours_from_cell(ws.cell(r, fri_c).value)
        if night_c:
            n += _hours_from_cell(ws.cell(r, night_c).value)
        if hourly_c:
            hourly += _hours_from_cell(ws.cell(r, hourly_c).value)
        if daily_c:
            try:
                daily += int(ws.cell(r, daily_c).value or 0)
            except (TypeError, ValueError):
                pass
        if sick_c:
            try:
                sick += int(ws.cell(r, sick_c).value or 0)
            except (TypeError, ValueError):
                pass

    real_ot = j - k if j >= k else 0.0
    if k <= j:
        real_tatil = m
    elif k <= j + m:
        real_tatil = j + m - k
    else:
        real_tatil = 0.0

    raw = {
        "work_days": work_days,
        "overtime": j,
        "kasr": k,
        "tatil": m,
        "friday": l,
        "night": n,
        "daily_leave": daily,
        "hourly_leave": hourly,
        "sick": sick,
        "real_overtime": real_ot,
        "real_tatil": real_tatil,
    }
    finalized = _finalize_sheet_totals(dict(raw))
    return {
        "work_days": finalized["work_days"],
        "real_overtime": finalized["summary_overtime"],
        "real_tatil": finalized["real_tatil"],
        "friday": finalized["friday"],
        "night": finalized["summary_night"],
        "daily_leave": finalized["daily_leave"],
        "hourly_leave": finalized["hourly_leave"],
        "sick": finalized["sick"],
        "total_kasr_cl": finalized["summary_kasr_cl"],
    }


def _summary_order(exp_wb) -> list[str]:
    ss = "کارکرد کلی1405.2"
    if ss not in exp_wb.sheetnames:
        return []
    ws = exp_wb[ss]
    names: list[str] = []
    for r in range(6, 40):
        name = ws.cell(r, 2).value
        if name:
            names.append(str(name).strip())
    return names


def _summary_by_name(wb, sheet_name: str) -> dict[str, dict[str, float]]:
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    out: dict[str, dict[str, float]] = {}
    for r in range(6, 50):
        name = ws.cell(r, 2).value
        if not name:
            continue
        out[str(name).strip()] = {
            "work_days": float(ws.cell(r, 3).value or 0),
            "real_overtime": _hours_from_cell(ws.cell(r, 4).value),
            "real_tatil": _hours_from_cell(ws.cell(r, 5).value),
            "friday": _hours_from_cell(ws.cell(r, 6).value),
            "night": _hours_from_cell(ws.cell(r, 7).value),
            "daily_leave": float(ws.cell(r, 8).value or 0),
            "hourly_leave": _hours_from_cell(ws.cell(r, 9).value),
            "sick": float(ws.cell(r, 10).value or 0),
            "total_kasr_cl": _hours_from_cell(ws.cell(r, 11).value),
        }
    return out


def diff_workbooks(got_path: Path, exp_path: Path) -> list[str]:
    got_wb = load_workbook(got_path, data_only=True)
    exp_wb = load_workbook(exp_path, data_only=True)
    errors: list[str] = []

    for emp in exp_wb.sheetnames:
        if emp.startswith("کارکرد کلی"):
            continue
        if emp not in got_wb.sheetnames:
            errors.append(f"missing sheet: {emp}")
            continue
        gws, ews = got_wb[emp], exp_wb[emp]

        gh, gc = _sheet_headers(gws)
        eh, ec = _sheet_headers(ews)
        if not gh or not eh:
            continue
        gdates = {}
        for r in range(gh + 1, gws.max_row + 1):
            d = str(gws.cell(r, gc["تاریخ"]).value or "").strip()
            if d.startswith("1405"):
                gdates[d] = r
        compare_cols = [
            c
            for c in ec
            if c.replace("\n", " ")
            in {
                "روز",
                "کارکرد",
                "کارکرد موظفی",
                "کارکرد موظف پس از کسر مرخصی",
                "اضافه کار",
                "کسرکار",
                "تعطیل کاری",
                "جمعه کاری",
                "شب کاری",
                "مرخصی",
                "نوع مرخصی",
            }
        ]
        for r in range(eh + 1, ews.max_row + 1):
            d = str(ews.cell(r, ec["تاریخ"]).value or "").strip()
            if not d.startswith("1405"):
                continue
            if d not in gdates:
                errors.append(f"{emp} {d}: missing date row")
                continue
            gr = gdates[d]
            for col in compare_cols:
                ecol = ec[col]
                gcol = None
                for k, v in gc.items():
                    if k.replace("\n", " ") == col.replace("\n", " "):
                        gcol = v
                        break
                if not gcol:
                    continue
                gv, ev = _norm(gws.cell(gr, gcol).value), _norm(ews.cell(r, ecol).value)
                ok = gv == ev or (
                    isinstance(gv, float) and isinstance(ev, float) and abs(gv - ev) < 0.03
                )
                if not ok:
                    errors.append(f"{emp} {d} {col}: got={gv!r} exp={ev!r}")

    ss = "کارکرد کلی1405.2"
    if ss not in got_wb.sheetnames:
        errors.append("missing summary sheet in output")
    else:
        got_by_name = _summary_by_name(got_wb, ss)
        exp_by_name = _summary_by_name(exp_wb, ss)
        for emp, exp_totals in exp_by_name.items():
            if emp not in got_by_name:
                errors.append(f"summary missing employee: {emp}")
                continue
            got_totals = got_by_name[emp]
            for key in exp_totals:
                gv, ev = got_totals.get(key), exp_totals[key]
                ok = gv == ev or (
                    isinstance(gv, float)
                    and isinstance(ev, float)
                    and abs(gv - ev) < 0.05
                )
                if not ok:
                    errors.append(f"summary {emp} {key}: got={gv!r} exp={ev!r}")

    got_wb.close()
    exp_wb.close()
    return errors


def main() -> int:
    if not RAW.is_file():
        print(f"FAIL: raw input missing: {RAW}")
        return 1
    if not EXPECTED.is_file():
        print(f"FAIL: expected missing: {EXPECTED}")
        return 1

    out_dir = Path("/tmp/karkard-verify")
    out_dir.mkdir(exist_ok=True)
    out = process_karkard_workbook(
        RAW,
        out_dir,
        jalali_year=1405,
    )
    print(f"output: {out}")

    errors = diff_workbooks(out, EXPECTED)
    if errors:
        print(f"FAIL: {len(errors)} differences")
        for line in errors[:30]:
            print(" ", line)
        if len(errors) > 30:
            print(f"  ... and {len(errors) - 30} more")
        return 1

    print("PASS: output matches expected reference")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
