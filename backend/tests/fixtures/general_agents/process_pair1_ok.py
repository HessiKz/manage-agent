from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook, load_workbook

def main(input_path: Path, output_dir: Path, *, agent_id: str, args: dict) -> Path:
    wb_in = load_workbook(input_path, data_only=True, read_only=True)
    out = Workbook()
    out.remove(out.active)
    summary = []
    for sn in wb_in.sheetnames:
        ws_in = wb_in[sn]
        rows = list(ws_in.iter_rows(values_only=True))
        ws = out.create_sheet(sn)
        if not rows:
            continue
        header = list(rows[0]) + ["total"]
        ws.append(header)
        line_count = 0
        grand = 0
        for r in rows[1:]:
            if not r or r[0] is None:
                continue
            qty = float(r[2] or 0)
            price = float(r[3] or 0)
            total = qty * price
            ws.append([r[0], r[1], qty, price, total])
            line_count += 1
            grand += total
        ws.append([None, "SUM", sum(float(r[2] or 0) for r in rows[1:] if r and r[0] is not None), None, grand])
        summary.append((sn, line_count, grand))
    ws = out.create_sheet("Summary")
    ws.append(["dept", "line_count", "grand_total"])
    total_lines = 0
    total_grand = 0
    for sn, lc, g in summary:
        ws.append([sn, lc, g])
        total_lines += lc
        total_grand += g
    ws.append(["ALL", total_lines, total_grand])
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"output_{input_path.stem}.xlsx"
    out.save(out_path)
    return out_path
