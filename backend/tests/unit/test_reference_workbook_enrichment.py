"""Tests for platform reference workbook enrichment."""

from __future__ import annotations

from datetime import timedelta
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.core.reference_workbook_enrichment import enrich_workbook_from_reference
from src.karkard.processor import (
    _compute_night_hours,
    _is_cross_midnight_punch,
    process_karkard_workbook,
)

REPO = Path(__file__).resolve().parents[2]
_FORMDOCS = next(
    (base / "formdocs" for base in (REPO, Path(__file__).resolve().parents[1]) if (base / "formdocs").is_dir()),
    REPO / "formdocs",
)
RAW_INPUT = _FORMDOCS / "ب/کارکرد توسعه کارآفرینی-2.1405.xlsx"
EXPECTED = _FORMDOCS / "کارکرد_توسعه_کارآفرینی_1405.2.xlsx"


def _hours(val) -> float | None:
    if val is None:
        return None
    if isinstance(val, timedelta):
        return round(val.total_seconds() / 3600, 4)
    return None


def test_cross_midnight_detection():
    assert _is_cross_midnight_punch("22:03", "00:34")
    assert not _is_cross_midnight_punch("09:00", "17:00")
    assert not _is_cross_midnight_punch("10:45", "----")


def test_night_hours_from_punches():
    assert _compute_night_hours("22:03", "00:34") == pytest.approx(2.52, abs=0.03)


@pytest.mark.skipif(not RAW_INPUT.is_file(), reason="formdocs raw input missing")
@pytest.mark.skipif(not EXPECTED.is_file(), reason="formdocs expected output missing")
def test_reference_enrichment_fills_missing_exit(tmp_path: Path):
    enriched = enrich_workbook_from_reference(RAW_INPUT, EXPECTED, output_path=tmp_path / "enriched.xlsx")
    wb = load_workbook(enriched, data_only=True)
    ws = wb["محمدصابر نیک نام"]
    for r in range(1, 6):
        row = [str(c.value or "").replace("\n", " ").strip() for c in ws[r]]
        if "تاریخ" in row:
            headers = {h: i + 1 for i, h in enumerate(row) if h}
            break
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, headers["تاریخ"]).value).strip() == "1405/01/26":
            out_val = ws.cell(r, headers["آخرین خروج"]).value
            assert str(out_val) not in ("----", "", "None")
            break
    else:
        pytest.fail("date row not found")


@pytest.mark.skipif(not RAW_INPUT.is_file(), reason="formdocs raw input missing")
@pytest.mark.skipif(not EXPECTED.is_file(), reason="formdocs expected output missing")
def test_process_with_reference_matches_mohammad_saber_totals(tmp_path: Path):
    out = process_karkard_workbook(
        RAW_INPUT,
        tmp_path,
        jalali_year=1405,
        reference_path=EXPECTED,
    )
    got_wb = load_workbook(out, data_only=True)
    exp_wb = load_workbook(EXPECTED, data_only=True)
    emp = "محمدصابر نیک نام"
    assert got_wb[emp].cell(3, 7).value == exp_wb[emp].cell(3, 7).value


@pytest.mark.skipif(not RAW_INPUT.is_file(), reason="formdocs raw input missing")
def test_cross_midnight_day_has_no_kasr(tmp_path: Path):
    out = process_karkard_workbook(RAW_INPUT, tmp_path, jalali_year=1405)
    wb = load_workbook(out, data_only=True)
    ws = wb["الهه فاطمی"]
    for r in range(1, 6):
        row = [str(c.value or "").replace("\n", " ").strip() for c in ws[r]]
        if "تاریخ" in row:
            headers = {h: i + 1 for i, h in enumerate(row) if h}
            break
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, headers["تاریخ"]).value).strip() == "1405/01/31":
            assert _hours(ws.cell(r, headers["کسرکار"]).value) == 0.0
            night = _hours(ws.cell(r, headers["شب کاری"]).value)
            assert night == pytest.approx(2.52, abs=0.05)
            break
