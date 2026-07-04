"""Output-sample detection on disk paths with uuid prefixes."""

from __future__ import annotations

from types import SimpleNamespace

from src.core.agent_file_roles import is_output_sample_file
from src.karkard.input_selection import is_runtime_karkard_candidate, pick_runtime_karkard_file


def test_output_sample_detected_in_storage_path():
    name = "deadbeef_output-sample__کارکرد_توسعه_کارآفرینی_1405.2.xlsx"
    assert is_output_sample_file(name)
    assert not is_runtime_karkard_candidate("کارکرد_توسعه_کارآفرینی_1405.2.xlsx", name)


def test_pick_runtime_prefers_raw_over_sample(tmp_path):
    from shutil import copy2

    from tests.test_karkard_locked_input import RAW_INPUT, SAMPLE_OUTPUT

    if not RAW_INPUT.is_file() or not SAMPLE_OUTPUT.is_file():
        import pytest

        pytest.skip("formdocs missing")

    agent_dir = tmp_path / "var" / "agent_files" / "a"
    agent_dir.mkdir(parents=True)
    sample_path = agent_dir / "uuid_output-sample__expected.xlsx"
    raw_path = agent_dir / "uuid_کارکرد توسعه کارآفرینی-2.1405.xlsx"
    copy2(SAMPLE_OUTPUT, sample_path)
    copy2(RAW_INPUT, raw_path)

    sample = SimpleNamespace(
        filename="output-sample__expected.xlsx",
        storage_path=str(sample_path),
    )
    raw = SimpleNamespace(
        filename="کارکرد توسعه کارآفرینی-2.1405.xlsx",
        storage_path=str(raw_path),
    )
    assert pick_runtime_karkard_file([sample, raw]) is raw


def test_pick_runtime_does_not_fallback_to_older_raw(tmp_path):
    """Newest upload wins; stale raw files must not be re-processed silently."""
    from shutil import copy2

    from tests.test_karkard_locked_input import RAW_INPUT, SAMPLE_OUTPUT

    if not RAW_INPUT.is_file() or not SAMPLE_OUTPUT.is_file():
        import pytest

        pytest.skip("formdocs missing")

    agent_dir = tmp_path / "var" / "agent_files" / "a"
    agent_dir.mkdir(parents=True)
    older_raw = agent_dir / "uuid_old_کارکرد توسعه کارآفرینی-2.1405.xlsx"
    newer_bad = agent_dir / "uuid_new_not_raw.xlsx"
    copy2(RAW_INPUT, older_raw)
    copy2(SAMPLE_OUTPUT, newer_bad)

    rows = [
        SimpleNamespace(filename="not_raw.xlsx", storage_path=str(newer_bad)),
        SimpleNamespace(filename="کارکرد توسعه کارآفرینی-2.1405.xlsx", storage_path=str(older_raw)),
    ]
    assert pick_runtime_karkard_file(rows) is None
    from shutil import copy2

    from openpyxl import load_workbook

    from src.karkard.input_selection import workbook_looks_like_raw_karkard
    from tests.test_karkard_locked_input import RAW_INPUT

    if not RAW_INPUT.is_file():
        import pytest

        pytest.skip("formdocs missing")

    path = tmp_path / "raw.xlsx"
    copy2(RAW_INPUT, path)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        assert wb.active.max_row is None or wb.active.max_row >= 2
    finally:
        wb.close()
    assert workbook_looks_like_raw_karkard(path) is True


def test_load_karkard_workbook_recovers_from_broken_styles(tmp_path):
    from unittest.mock import patch

    from openpyxl import Workbook, load_workbook

    from src.karkard.processor import _load_karkard_workbook
    from tests.test_karkard_locked_input import RAW_INPUT

    if not RAW_INPUT.is_file():
        import pytest

        pytest.skip("formdocs missing")

    path = tmp_path / "raw.xlsx"
    wb = load_workbook(RAW_INPUT)
    wb.save(path)
    wb.close()

    real_load = load_workbook

    def fake_load(p, *args, **kwargs):
        if kwargs.get("read_only"):
            return real_load(p, *args, **kwargs)
        raise IndexError("style_id")

    with patch("src.karkard.processor.load_workbook", side_effect=fake_load):
        recovered = _load_karkard_workbook(path)
    assert recovered.sheetnames
    ws = recovered[recovered.sheetnames[0]]
    assert _find_header_row_from_wb(ws) is not None


def _find_header_row_from_wb(ws):
    from src.karkard.processor import _find_header_row

    return _find_header_row(ws)
