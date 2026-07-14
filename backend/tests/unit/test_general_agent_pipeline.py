"""General (domain-neutral) file-agent pipeline: schema, multi-sheet verify, sandbox, roles."""

from __future__ import annotations

import shutil
from pathlib import Path
from types import SimpleNamespace
from uuid import uuid4

import pytest

from src.core.agent_file_roles import (
    ROLE_OUTPUT_SAMPLE,
    ROLE_RUNTIME,
    agent_file_role,
    pair_id_from_filename,
)
from src.services.agent_script_service import (
    AgentScriptService,
    run_agent_script_file,
    verify_file_matches,
)
from src.services.io_schema_service import build_io_schema_pair, schema_for_path

FIXTURES = Path(__file__).resolve().parents[1] / "fixtures" / "general_agents"


class _Scalars:
    def __init__(self, rows):
        self.rows = rows

    def all(self):
        return self.rows


class _Result:
    def __init__(self, rows):
        self.rows = rows

    def scalars(self):
        return _Scalars(self.rows)


class _Db:
    def __init__(self, rows=()):
        self.rows = list(rows)

    async def execute(self, *_args, **_kwargs):
        return _Result(self.rows)

    async def flush(self):
        pass


def _agent(**kw):
    return SimpleNamespace(
        id=kw.get("id", uuid4()),
        slug=kw.get("slug", "gen_agent"),
        name=kw.get("name", "General Agent"),
        description=kw.get("description", "multi-sheet transform"),
        system_prompt=kw.get("system_prompt", ""),
        kind=kw.get("kind", SimpleNamespace(value="worker")),
        capabilities=kw.get("capabilities", {"file_upload_enabled": True}),
        config_json=kw.get("config_json", {}),
        model_name=kw.get("model_name", None),
    )


def _stage(root: Path, src: Path, name: str, *, role: str | None = None, pair_id: str | None = None):
    dest = root / f"{uuid4().hex}_{name}"
    shutil.copy2(src, dest)
    return SimpleNamespace(
        filename=name,
        storage_path=str(dest),
        created_at=None,
        role=role or agent_file_role(name),
        pair_id=pair_id or pair_id_from_filename(name),
    )


def test_schema_multi_sheet_pair1():
    schema = schema_for_path(FIXTURES / "pair1_input.xlsx")
    assert schema["format"] == "xlsx"
    assert set(schema["sheet_names"]) == {"DeptA", "DeptB"}
    assert schema["sheet_row_counts"]["DeptA"] == 3
    out = schema_for_path(FIXTURES / "pair1_output.xlsx")
    assert "Summary" in out["sheet_names"]
    assert out["sheet_row_counts"]["Summary"] == 4


def test_verify_multi_sheet_ok_and_header_fail(tmp_path):
    actual = tmp_path / "a.xlsx"
    expected = FIXTURES / "pair1_output.xlsx"
    shutil.copy2(expected, actual)
    verify_file_matches(actual, expected)

    # Header mismatch is a hard structural fail
    from openpyxl import load_workbook

    wb = load_workbook(actual)
    ws = wb["DeptA"]
    ws.cell(1, 1).value = "WRONG"
    wb.save(actual)
    with pytest.raises(ValueError, match="Header"):
        verify_file_matches(actual, expected)


def test_verify_sheet_name_mismatch(tmp_path):
    from openpyxl import Workbook

    bad = tmp_path / "bad.xlsx"
    wb = Workbook()
    wb.active.title = "OnlyOne"
    wb.save(bad)
    with pytest.raises(ValueError, match="Sheet name"):
        verify_file_matches(bad, FIXTURES / "pair1_output.xlsx")


def test_verify_csv_pair():
    verify_file_matches(FIXTURES / "orders_output.csv", FIXTURES / "orders_output.csv")
    with pytest.raises(ValueError):
        verify_file_matches(FIXTURES / "orders_input.csv", FIXTURES / "orders_output.csv")


def test_no_domain_footer_trust_by_default(tmp_path):
    """Footer KEEP must not soft-pass general agents when values differ."""
    from openpyxl import Workbook

    expected = tmp_path / "exp.xlsx"
    actual = tmp_path / "act.xlsx"
    for path, val in ((expected, 100), (actual, 1)):
        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        ws.append(["label", "amount"])
        ws.append(["جمع", val])
        wb.save(path)
    with pytest.raises(ValueError, match="accuracy|Cell mismatch|Header"):
        verify_file_matches(actual, expected, min_accuracy=0.99, trust_domain_footers=False)


def test_pair_id_from_filename():
    assert pair_id_from_filename("pair1_input.xlsx") == "pair1"
    assert pair_id_from_filename("output-sample__pair2_out.xlsx") == "pair2"


@pytest.mark.anyio
async def test_sample_pairs_multi(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    agent = _agent()
    root = tmp_path / "var" / "agent_files" / str(agent.id)
    root.mkdir(parents=True)
    rows = [
        _stage(root, FIXTURES / "pair1_input.xlsx", "pair1_input.xlsx", pair_id="pair1"),
        _stage(
            root,
            FIXTURES / "pair1_output.xlsx",
            "output-sample__pair1_output.xlsx",
            role=ROLE_OUTPUT_SAMPLE,
            pair_id="pair1",
        ),
        _stage(root, FIXTURES / "pair2_input.xlsx", "pair2_input.xlsx", pair_id="pair2"),
        _stage(
            root,
            FIXTURES / "pair2_output.xlsx",
            "output-sample__pair2_output.xlsx",
            role=ROLE_OUTPUT_SAMPLE,
            pair_id="pair2",
        ),
    ]
    pairs = await AgentScriptService(_Db(rows))._sample_pairs(agent)
    assert len(pairs) == 2


def test_sandbox_runs_ok_script(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    agent_id = uuid4()
    root = tmp_path / "var" / "agent_files" / str(agent_id)
    scripts = root / "scripts"
    scripts.mkdir(parents=True)
    shutil.copy2(FIXTURES / "process_pair1_ok.py", scripts / "process_gen_agent.py")
    inp = root / "in.xlsx"
    shutil.copy2(FIXTURES / "pair1_input.xlsx", inp)
    agent = _agent(
        id=agent_id,
        config_json={
            "workspace_script": {
                "needed": True,
                "slug": "process_gen_agent",
                "path": "scripts/process_gen_agent.py",
            }
        },
    )
    out = run_agent_script_file(agent, inp)
    assert out.is_file()
    verify_file_matches(out, FIXTURES / "pair1_output.xlsx")


def test_sandbox_timeout(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    monkeypatch.setenv("MA_SCRIPT_TIMEOUT_S", "1")
    agent_id = uuid4()
    root = tmp_path / "var" / "agent_files" / str(agent_id)
    scripts = root / "scripts"
    scripts.mkdir(parents=True)
    (scripts / "hang.py").write_text(
        "from pathlib import Path\nimport time\n"
        "def main(input_path, output_dir, *, agent_id, args):\n"
        "    time.sleep(30)\n"
        "    return Path(output_dir) / 'x.txt'\n",
        encoding="utf-8",
    )
    inp = root / "in.txt"
    inp.write_text("x", encoding="utf-8")
    agent = _agent(
        id=agent_id,
        config_json={
            "workspace_script": {
                "needed": True,
                "slug": "hang",
                "path": "scripts/hang.py",
            }
        },
    )
    with pytest.raises((TimeoutError, RuntimeError)):
        run_agent_script_file(agent, inp)


def test_sandbox_rejects_path_escape_input(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    agent_id = uuid4()
    root = tmp_path / "var" / "agent_files" / str(agent_id)
    scripts = root / "scripts"
    scripts.mkdir(parents=True)
    (scripts / "x.py").write_text(
        "from pathlib import Path\nfrom shutil import copy2\n"
        "def main(input_path, output_dir, *, agent_id, args):\n"
        "    output_dir.mkdir(parents=True, exist_ok=True)\n"
        "    out = Path(output_dir) / 'o.txt'\n"
        "    copy2(input_path, out)\n"
        "    return out\n",
        encoding="utf-8",
    )
    outside = tmp_path / "outside.txt"
    outside.write_text("nope", encoding="utf-8")
    agent = _agent(
        id=agent_id,
        config_json={
            "workspace_script": {
                "needed": True,
                "slug": "x",
                "path": "scripts/x.py",
            }
        },
    )
    with pytest.raises(PermissionError):
        run_agent_script_file(agent, outside)


def test_build_io_schema_pair():
    pack = build_io_schema_pair(FIXTURES / "pair1_input.xlsx", FIXTURES / "pair1_output.xlsx")
    assert pack["input"]["sheet_names"]
    assert pack["output"]["sheet_names"]


@pytest.mark.anyio
async def test_csv_verify_via_service_copy_stub(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    agent = _agent()
    root = tmp_path / "var" / "agent_files" / str(agent.id)
    root.mkdir(parents=True)
    # identical input/output so copy-stub verifies
    content = (FIXTURES / "orders_output.csv").read_text(encoding="utf-8")
    rows = [
        _stage(root, FIXTURES / "orders_output.csv", "input.csv", role=ROLE_RUNTIME),
        _stage(
            root,
            FIXTURES / "orders_output.csv",
            "output-sample__expected.csv",
            role=ROLE_OUTPUT_SAMPLE,
        ),
    ]
    # rewrite in case stage names differ
    Path(rows[0].storage_path).write_text(content, encoding="utf-8")
    Path(rows[1].storage_path).write_text(content, encoding="utf-8")
    meta = await AgentScriptService(_Db(rows)).verify(agent, use_llm=False)
    assert meta.get("verified_at")


def test_role_helpers():
    assert agent_file_role("output-sample__x.xlsx") == ROLE_OUTPUT_SAMPLE
    assert agent_file_role("x.xlsx", role="instruction") == "instruction"
