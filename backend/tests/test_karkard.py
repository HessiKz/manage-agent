"""کارکرد spreadsheet processor and HR agent action."""

from __future__ import annotations

import uuid
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from src.karkard.output import KARKARD_OUTPUT_DIR
from src.karkard.processor import process_karkard_workbook
from src.main import app

FIXTURE_XLSX = next(
    (
        base / "formdocs/ب/کارکرد توسعه کارآفرینی-2.1405.xlsx"
        for base in (
            Path(__file__).resolve().parents[2],
            Path(__file__).resolve().parents[1],
        )
        if (base / "formdocs/ب/کارکرد توسعه کارآفرینی-2.1405.xlsx").is_file()
    ),
    Path(__file__).resolve().parents[2] / "formdocs/ب/کارکرد توسعه کارآفرینی-2.1405.xlsx",
)
FIXTURE_FALLBACK = Path(__file__).resolve().parent / "fixtures/karkard_sample.xlsx"


def _fixture_xlsx() -> Path:
    if FIXTURE_XLSX.is_file():
        return FIXTURE_XLSX
    return FIXTURE_FALLBACK


@pytest.fixture(scope="module")
def client():
    with TestClient(app) as c:
        yield c


@pytest.fixture(scope="module")
def auth_headers(client: TestClient):
    login = client.post(
        "/api/v1/auth/login",
        json={"email": "admin@example.com", "password": "admin123"},
    )
    assert login.status_code == 200
    return {"Authorization": f"Bearer {login.json()['access_token']}"}


@pytest.mark.skipif(not _fixture_xlsx().is_file(), reason="formdocs fixture missing")
def test_process_karkard_workbook_columns(tmp_path: Path):
    out_dir = tmp_path / "out"
    out_path = process_karkard_workbook(_fixture_xlsx(), out_dir)
    assert out_path.is_file()

    wb = load_workbook(out_path, data_only=True)
    header_row = None
    headers: list[str] = []
    for name in wb.sheetnames:
        if name.startswith("کارکرد کلی"):
            continue
        ws = wb[name]
        for r in range(1, 6):
            vals = [str(c.value or "") for c in ws[r]]
            if "تاریخ" in vals:
                header_row = r
                headers = vals
                break
        if header_row is not None:
            break
    assert header_row is not None
    assert "کارکرد موظف پس از کسر مرخصی" in headers
    assert any("کسرکار" in h or "تاخیر" in h for h in headers)


def test_karkard_tool_not_registered():
    import src.agents_lib.custom_tools  # noqa: F401
    from src.agents_lib.tool_registry import ToolRegistry

    assert "karkard_process" not in ToolRegistry.list_slugs()
    assert "run_agent_script" in ToolRegistry.list_slugs()


@pytest.mark.skip(reason="karkard_process removed; processing is via run_agent_script + pinned script")
@pytest.mark.skipif(not _fixture_xlsx().is_file(), reason="formdocs fixture missing")
def test_karkard_agent_action_flow(
    client: TestClient,
    auth_headers: dict,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    from tests.helpers.mock_react import mock_run_react_agent

    monkeypatch.setattr(
        "src.services.orchestrator_service.run_react_agent",
        mock_run_react_agent,
    )
    monkeypatch.setattr("src.karkard.output.KARKARD_OUTPUT_DIR", tmp_path)
    suffix = uuid.uuid4().hex[:8]
    create = client.post(
        "/api/v1/agents",
        headers=auth_headers,
        json={
            "name": f"Karkard Test {suffix}",
            "slug": f"karkard-test-{suffix}",
            "kind": "worker",
            "capabilities": {
                "chat_enabled": True,
                "file_upload_enabled": True,
                "actions_enabled": True,
                "templates_enabled": True,
            },
            "file_policy": {
                "min_files": 1,
                "max_files": 5,
                "max_file_size_mb": 25,
                "max_total_size_mb": 100,
                "allowed_mime_types": [
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ],
                "allowed_extensions": [".xlsx"],
                "require_files_to_invoke": False,
                "auto_ingest_to_rag": False,
            },
            "tool_names": ["run_agent_script"],
            "actions": [
                {
                    "slug": "process_karkard",
                    "label": "محاسبه",
                    "prompt_template": "پردازش کارکرد",
                    "tool_chain": ["run_agent_script"],
                    "input_schema": {},
                }
            ],
        },
    )
    assert create.status_code == 201, create.text
    agent_id = create.json()["id"]

    with _fixture_xlsx().open("rb") as fh:
        upload = client.post(
            f"/api/v1/agents/{agent_id}/files",
            headers=auth_headers,
            files={
                "file": (
                    "raw.xlsx",
                    fh,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert upload.status_code in (200, 201), upload.text

    run = client.post(
        f"/api/v1/agents/{agent_id}/actions/process_karkard/run",
        headers=auth_headers,
        json={"variables": {"jalali_year": 1405}},
    )
    assert run.status_code == 200, run.text
    body = run.json()
    output = body["output"]
    assert "download_path" in output or "karkard" in output or "/workspace/" in output
    assert "/api/v1/agents/" in output and "/workspace/" in output

    agent_dir = Path("var/agent_files") / agent_id
    files = list(agent_dir.glob("karkard-[0-9a-f][0-9a-f][0-9a-f][0-9a-f][0-9a-f][0-9a-f][0-9a-f][0-9a-f].xlsx")) if agent_dir.is_dir() else []
    if not files:
        files = list(tmp_path.glob("karkard-[0-9a-f][0-9a-f][0-9a-f][0-9a-f][0-9a-f][0-9a-f][0-9a-f][0-9a-f].xlsx"))
    assert files, "expected processed xlsx beside upload or in output dir"
    assert files[-1].stat().st_size > 1000


def test_catalog_karkard_agent_after_seed(client: TestClient, auth_headers: dict):
    r = client.get("/api/v1/agents/by-slug/example-karkard", headers=auth_headers)
    if r.status_code == 404:
        pytest.skip("example-karkard not seeded")
    assert r.status_code == 200
    data = r.json()
    assert data["kind"] == "worker"
    assert data["capabilities"]["file_upload_enabled"] is True
    assert "run_agent_script" in data["tool_names"]
