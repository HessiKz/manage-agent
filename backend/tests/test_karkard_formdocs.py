"""Regression: real formdocs input must match HR example output."""

from __future__ import annotations

from datetime import timedelta
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.karkard.processor import process_karkard_workbook

REPO = Path(__file__).resolve().parents[2]
_FORMDOCS = next(
    (base / "formdocs" for base in (REPO, Path(__file__).resolve().parents[1]) if (base / "formdocs").is_dir()),
    REPO / "formdocs",
)
RAW_INPUT = _FORMDOCS / "ب/کارکرد توسعه کارآفرینی-2.1405.xlsx"
EXPECTED = _FORMDOCS / "کارکرد_توسعه_کارآفرینی_1405.2.xlsx"


def _hours(val) -> float | None:
    if val is None:
        return None
    if isinstance(val, timedelta):
        return round(val.total_seconds() / 3600, 4)
    if isinstance(val, str) and ":" in val:
        parts = val.split(":")
        try:
            if len(parts) == 3:
                h, m, s = map(int, parts)
                return round(h + m / 60 + s / 3600, 4)
            if len(parts) == 2:
                h, m = map(int, parts)
                return round(h + m / 60, 4)
        except ValueError:
            return None
    return None


def _sheet_map(wb, name: str) -> dict[str, dict]:
    ws = wb[name]
    header_row = None
    headers: dict[str, int] = {}
    for r in range(1, 6):
        row = [str(c.value or "").strip() for c in ws[r]]
        if "تاریخ" in row:
            header_row = r
            for i, h in enumerate(row, start=1):
                if h:
                    headers[h.replace("\n", " ")] = i
            break
    assert header_row is not None
    out: dict[str, dict] = {}
    date_col = headers.get("تاریخ")
    for r in range(header_row + 1, ws.max_row + 1):
        raw_date = ws.cell(r, date_col).value
        if not raw_date:
            continue
        key = str(raw_date).strip()
        if not key.startswith("1405"):
            continue
        out[key] = {
            "in": str(ws.cell(r, headers["اولین ورود"]).value or "")
            if headers.get("اولین ورود")
            else "",
            "out": str(ws.cell(r, headers["آخرین خروج"]).value or "")
            if headers.get("آخرین خروج")
            else "",
            "day": ws.cell(r, headers.get("روز", 0)).value if headers.get("روز") else None,
            "work": _hours(ws.cell(r, headers["کارکرد"]).value) if "کارکرد" in headers else None,
            "after": _hours(
                ws.cell(
                    r,
                    headers.get("کارکرد موظف پس از کسر مرخصی")
                    or headers.get("کارکرد موظف\nپس از کسر مرخصی", 0),
                ).value
            )
            if headers.get("کارکرد موظف پس از کسر مرخصی")
            or headers.get("کارکرد موظف\nپس از کسر مرخصی")
            else None,
            "ot": _hours(ws.cell(r, headers.get("اضافه کار", 0)).value)
            if headers.get("اضافه کار")
            else None,
            "kasr": _hours(ws.cell(r, headers.get("کسرکار", 0)).value)
            if headers.get("کسرکار")
            else None,
            "tatil": _hours(
                ws.cell(r, headers.get("تعطیل کاری") or headers.get("تعطیلکاری", 0)).value
            )
            if headers.get("تعطیل کاری") or headers.get("تعطیلکاری")
            else None,
        }
    return out


def _raw_sheet_map(path: Path, name: str) -> dict[str, dict]:
    wb = load_workbook(path, data_only=True)
    return _sheet_map(wb, name)


@pytest.mark.skipif(not RAW_INPUT.is_file(), reason="formdocs raw input missing")
@pytest.mark.skipif(not EXPECTED.is_file(), reason="formdocs expected output missing")
def test_formdocs_matches_expected_on_identical_punch_rows(tmp_path: Path):
    out = process_karkard_workbook(RAW_INPUT, tmp_path, jalali_year=1405)
    got_wb = load_workbook(out, data_only=True)
    exp_wb = load_workbook(EXPECTED, data_only=True)
    raw_map = _raw_sheet_map(RAW_INPUT, "مهدی چشمه کبودی")

    got = _sheet_map(got_wb, "مهدی چشمه کبودی")
    exp = _sheet_map(exp_wb, "مهدی چشمه کبودی")
    assert set(got.keys()) == set(exp.keys())

    for date, raw_row in raw_map.items():
        if raw_row["in"] != exp[date]["in"] or raw_row["out"] != exp[date]["out"]:
            continue
        g, e = got[date], exp[date]
        assert g["day"] == e["day"], date
        assert g["work"] == pytest.approx(e["work"], abs=0.03), date
        assert g["after"] == pytest.approx(e["after"], abs=0.03), date
        assert g["ot"] == pytest.approx(e["ot"], abs=0.03), date
        assert g["kasr"] == pytest.approx(e["kasr"], abs=0.03), date
        assert g["tatil"] == pytest.approx(e["tatil"], abs=0.03), date

    thu = exp["1405/01/27"]
    assert thu["day"] == "پنجشنبه"
    assert got["1405/01/27"]["tatil"] == pytest.approx(thu["tatil"], abs=0.02)


@pytest.mark.skipif(not RAW_INPUT.is_file(), reason="formdocs raw input missing")
@pytest.mark.skipif(not EXPECTED.is_file(), reason="formdocs expected output missing")
def test_formdocs_summary_sheet_name(tmp_path: Path):
    out = process_karkard_workbook(RAW_INPUT, tmp_path, jalali_year=1405)
    wb = load_workbook(out, read_only=True)
    assert "کارکرد کلی1405.2" in wb.sheetnames
    wb.close()


@pytest.mark.skipif(not RAW_INPUT.is_file(), reason="formdocs raw input missing")
@pytest.mark.skipif(not EXPECTED.is_file(), reason="formdocs expected output missing")
def test_incomplete_punch_days_do_not_accumulate_kasr(tmp_path: Path):
    out = process_karkard_workbook(RAW_INPUT, tmp_path, jalali_year=1405)
    got_wb = load_workbook(out, data_only=True)
    got = _sheet_map(got_wb, "محمدصابر نیک نام")
    assert got["1405/01/26"]["kasr"] == 0.0
    assert got["1405/01/29"]["kasr"] == 0.0


@pytest.mark.skipif(not EXPECTED.is_file(), reason="formdocs expected output missing")
def test_rejects_already_processed_workbook(tmp_path: Path):
    with pytest.raises(ValueError, match="نمونه خروجی|پردازش"):
        process_karkard_workbook(EXPECTED, tmp_path, jalali_year=1405)
